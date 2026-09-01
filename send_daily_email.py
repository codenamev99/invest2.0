from __future__ import annotations

import html
import os
import smtplib
from datetime import date, datetime
from email.message import EmailMessage
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


RESULTS_PATH = Path(os.environ.get("RESULTS_FILE", "results.xlsx"))
HOW_IT_WORKS_SHEET_NAME = "How It Works"
RECENT_SPLITS_SHEET_NAME = "Recent Splits (90D)"
UPCOMING_IPOS_SHEET_NAME = "Upcoming IPOs (60D)"
UPCOMING_EARNINGS_SHEET_NAME = "Upcoming Earnings (14D)"
SIMULATION_SHEET_NAME = "Simulation"
AM_SIMULATION_SHEET_NAME = "AM Simulation"
PM_SIMULATION_SHEET_NAME = "PM Simulation"
COMMIT_SUMMARY_SHEET_NAME = "Commit Summary"
LEGACY_SIMULATION_SHEET_NAME = "Summary"
DAILY_RUNS_SHEET_NAME = "Daily Runs"
# Lower-cased and newline-stripped, matching how the sheet headers are normalized.
MARKET_CONDITION_HEADER = "spy - market condition"
PROTECTED_SHEETS = {
    "Single Tickers",
    HOW_IT_WORKS_SHEET_NAME,
    RECENT_SPLITS_SHEET_NAME,
    UPCOMING_IPOS_SHEET_NAME,
    UPCOMING_EARNINGS_SHEET_NAME,
    "Top 10 OHLC Tracking",
    SIMULATION_SHEET_NAME,
    AM_SIMULATION_SHEET_NAME,
    PM_SIMULATION_SHEET_NAME,
    COMMIT_SUMMARY_SHEET_NAME,
    LEGACY_SIMULATION_SHEET_NAME,
    "Investment Dashboard",
    DAILY_RUNS_SHEET_NAME,
}


def required_env(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise SystemExit(f"Missing required environment variable: {name}")
    return value


def parse_run_sheet_date(sheet_name: str) -> date | None:
    try:
        return datetime.strptime(sheet_name.strip(), "%d %b %Y").date()
    except ValueError:
        return None


def latest_run_sheet_name(sheet_names: list[str]) -> str | None:
    candidates = [
        (parsed, name)
        for name in sheet_names
        if name not in PROTECTED_SHEETS
        for parsed in [parse_run_sheet_date(name)]
        if parsed is not None
    ]
    if candidates:
        return max(candidates, key=lambda item: item[0])[1]

    for name in reversed(sheet_names):
        if name not in PROTECTED_SHEETS:
            return name
    return None


def format_value(value: Any, number_format: str = "") -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%b %-d, %Y")
    if isinstance(value, date):
        return value.strftime("%b %-d, %Y")
    if isinstance(value, float):
        if '0.00"%"' in number_format:
            return f"{value:.2f}%"
        if "0.00%" in number_format:
            return f"{value:.2%}"
        if "0%" in number_format and abs(value) <= 1:
            return f"{value:.0%}"
        if "$" in number_format:
            return f"${value:,.2f}"
        return f"{value:,.2f}"
    if isinstance(value, int):
        if '0.00"%"' in number_format:
            return f"{value:.2f}%"
        if "0.00%" in number_format:
            return f"{value:.2%}"
        if "0%" in number_format and abs(value) <= 1:
            return f"{value:.0%}"
        if "$" in number_format:
            return f"${value:,.2f}"
        return f"{value:,}"
    return str(value)


def html_table(headers: list[str], rows: list[list[str]]) -> str:
    if not rows:
        return "<p>No rows found.</p>"

    header_html = "".join(f"<th>{html.escape(header)}</th>" for header in headers)
    body_html = []
    for row in rows:
        cells = "".join(f"<td>{html.escape(cell)}</td>" for cell in row)
        body_html.append(f"<tr>{cells}</tr>")

    return f"""
    <table>
      <thead><tr>{header_html}</tr></thead>
      <tbody>{''.join(body_html)}</tbody>
    </table>
    """


def worksheet_table(wb, sheet_name: str, max_rows: int | None = None) -> str:
    if sheet_name not in wb.sheetnames:
        return f"<p>No {html.escape(sheet_name)} sheet was found.</p>"

    ws = wb[sheet_name]
    header_row = None
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
        if sum(value is not None for value in values) > 1:
            header_row = row_idx
            break
    if header_row is None:
        return f"<p>No rows found in {html.escape(sheet_name)}.</p>"

    headers = [
        str(ws.cell(row=header_row, column=col_idx).value or "").strip() or f"Column {col_idx}"
        for col_idx in range(1, ws.max_column + 1)
    ]
    rows: list[list[str]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
        if not any(values):
            continue
        rows.append(
            [
                format_value(ws.cell(row=row_idx, column=col_idx).value, ws.cell(row=row_idx, column=col_idx).number_format)
                for col_idx in range(1, ws.max_column + 1)
            ]
        )
        if max_rows is not None and len(rows) >= max_rows:
            break

    return html_table(headers, rows)


def ranked_stocks_table(wb) -> tuple[str, str]:
    if DAILY_RUNS_SHEET_NAME in wb.sheetnames:
        ws = wb[DAILY_RUNS_SHEET_NAME]
        headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
        try:
            date_col = headers.index("run date") + 1
            symbol_col = headers.index("symbol") + 1
        except ValueError:
            date_col, symbol_col = 1, 2
        company_col = next((i + 1 for i, value in enumerate(headers) if value == "company"), 3)
        rank_col = next((i + 1 for i, value in enumerate(headers) if value == "rank"), None)
        # Headers were renamed to plain English ("Closing Price", "Earnings Dates");
        # both spellings are matched so older workbooks still resolve.
        close_col = next(
            (
                i + 1
                for i, value in enumerate(headers)
                if value == "closing price" or ("close" in value and "$" in value)
            ),
            6,
        )
        earnings_cols = [i + 1 for i, value in enumerate(headers) if value.startswith("earnings")]
        next_earnings_col = (earnings_cols[-1] + 1) if earnings_cols else ws.max_column
        run_dates = [
            parse_run_date_value(ws.cell(row=row_idx, column=date_col).value)
            for row_idx in range(3, ws.max_row + 1)
        ]
        latest_date = max((value for value in run_dates if value is not None), default=None)
        selected_cols = [("Symbol", symbol_col), ("Company", company_col)]
        if rank_col:
            selected_cols.append(("Rank", rank_col))
        selected_cols.extend([("Close", close_col), ("Next Earnings", next_earnings_col)])
        rows: list[list[str]] = []
        for row_idx in range(3, ws.max_row + 1):
            if parse_run_date_value(ws.cell(row=row_idx, column=date_col).value) != latest_date:
                continue
            rows.append([
                format_value(ws.cell(row=row_idx, column=col_idx).value, ws.cell(row=row_idx, column=col_idx).number_format)
                for _, col_idx in selected_cols
            ])
            if len(rows) >= 10:
                break
        title_date = latest_date.strftime("%d %b %Y") if latest_date else "Latest"
        return f"Latest Ranked Stocks ({html.escape(title_date)})", html_table([label for label, _ in selected_cols], rows)

    sheet_name = latest_run_sheet_name(wb.sheetnames)
    if not sheet_name:
        return "Latest Ranked Stocks", "<p>No ranked stock sheet was found.</p>"

    ws = wb[sheet_name]
    rank_col = 3 if str(ws.cell(row=1, column=3).value or "").strip().lower() == "rank" else None
    selected_cols = [
        ("Symbol", 1),
        ("Company", 2),
    ]
    if rank_col:
        selected_cols.append(("Rank", rank_col))
    selected_cols.extend(
        [
            ("Close", 5 if rank_col else 4),
            ("Next Earnings", 27 if rank_col else 26),
        ]
    )

    rows: list[list[str]] = []
    for row_idx in range(3, ws.max_row + 1):
        symbol = ws.cell(row=row_idx, column=1).value
        if not symbol:
            continue
        rows.append(
            [
                format_value(ws.cell(row=row_idx, column=col_idx).value, ws.cell(row=row_idx, column=col_idx).number_format)
                for _, col_idx in selected_cols
            ]
        )
        if len(rows) >= 10:
            break

    headers = [label for label, _ in selected_cols]
    title = f"Latest Ranked Stocks ({html.escape(sheet_name)})"
    return title, html_table(headers, rows)


def parse_run_date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None


def upcoming_ipos_table(wb) -> str:
    return worksheet_table(wb, UPCOMING_IPOS_SHEET_NAME, max_rows=20)


def simulation_sheet_totals_html(wb, sheet_names: tuple[str, ...], label: str) -> str:
    sheet_name = next((name for name in sheet_names if name in wb.sheetnames), None)
    if sheet_name is None:
        return f"<p>{html.escape(label)} totals: Simulation sheet not found.</p>"

    ws = wb[sheet_name]
    normalized_headers = [
        str(ws.cell(row=1, column=col_idx).value or "").replace("\n", " ").strip().lower()
        for col_idx in range(1, ws.max_column + 1)
    ]
    try:
        result_currency_col = normalized_headers.index("result $") + 1
        result_pct_col = normalized_headers.index("result %") + 1
    except ValueError:
        return f"<p>{html.escape(label)} totals: Result columns not found.</p>"

    # The sheet totals are SUMIF(..., "Good*", ...) over the market-condition
    # column, so blocked, excluded, ignored and pending rows are left out. The
    # fallback below has to apply the same filter or the email reports trades the
    # workbook does not count. Sheets written without that column (the legacy
    # Summary tab) total with a plain SUM, which is what a missing column means.
    condition_col = (
        normalized_headers.index(MARKET_CONDITION_HEADER) + 1
        if MARKET_CONDITION_HEADER in normalized_headers
        else None
    )
    total_label_row = None
    for row_idx in range(1, ws.max_row + 1):
        labels = {
            str(ws.cell(row=row_idx, column=result_currency_col).value or "").strip().upper(),
            str(ws.cell(row=row_idx, column=result_pct_col).value or "").strip().upper(),
        }
        if "TOTAL" in labels:
            total_label_row = row_idx
            break

    if total_label_row is None:
        return f"<p>{html.escape(label)} totals: TOTAL row not found.</p>"

    total_formula_row = total_label_row + 1
    dollar_value = ws.cell(row=total_formula_row, column=result_currency_col).value
    percent_value = ws.cell(row=total_formula_row, column=result_pct_col).value

    def counted_total(column: int) -> float:
        """Sum one result column the way the sheet's own TOTAL formula does."""
        total = 0.0
        for row_idx in range(2, total_label_row):
            if condition_col is not None:
                condition = str(ws.cell(row=row_idx, column=condition_col).value or "")
                if not condition.strip().lower().startswith("good"):
                    continue
            value = ws.cell(row=row_idx, column=column).value
            if isinstance(value, (int, float)):
                total += float(value)
        return total

    # openpyxl writes formulas but never evaluates them, so a workbook this
    # pipeline just wrote has no cached result for the TOTAL cells and these read
    # back as None. The recomputed totals are the normal path, not a rare
    # fallback; a real number only appears here once Excel has saved the file.
    if not isinstance(dollar_value, (int, float)):
        dollar_value = counted_total(result_currency_col)
    if not isinstance(percent_value, (int, float)):
        percent_value = counted_total(result_pct_col)

    dollar_total = format_value(dollar_value, '"$"#,##0.00')
    percent_total = format_value(percent_value, "0.00%")
    return f"<p><strong>{html.escape(label)} totals:</strong> {html.escape(dollar_total)} and {html.escape(percent_total)}.</p>"


def simulation_totals_html(wb) -> str:
    regular_totals = simulation_sheet_totals_html(
        wb,
        (SIMULATION_SHEET_NAME, LEGACY_SIMULATION_SHEET_NAME),
        "Simulated Portfolio",
    )
    am_totals = simulation_sheet_totals_html(
        wb,
        (AM_SIMULATION_SHEET_NAME,),
        "AM Simulated Portfolio",
    )
    pm_totals = simulation_sheet_totals_html(
        wb,
        (PM_SIMULATION_SHEET_NAME,),
        "PM Simulated Portfolio",
    )
    return regular_totals + am_totals + pm_totals


def build_email_html() -> tuple[str, str]:
    if not RESULTS_PATH.exists():
        raise SystemExit(f"Results workbook not found: {RESULTS_PATH}")

    wb = load_workbook(RESULTS_PATH, data_only=True)
    ranked_title, ranked_html = ranked_stocks_table(wb)
    simulation_totals = simulation_totals_html(wb)
    ipo_html = upcoming_ipos_table(wb)
    run_date = datetime.now().strftime("%b %-d, %Y")
    subject_prefix = os.environ.get("EMAIL_SUBJECT_PREFIX", "Daily Screener").strip() or "Daily Screener"
    subject = f"{subject_prefix} - {run_date}"

    body = f"""
    <!doctype html>
    <html>
      <head>
        <style>
          body {{ font-family: Arial, sans-serif; color: #111; }}
          table {{ border-collapse: collapse; margin-bottom: 24px; }}
          th, td {{ border: 1px solid #bbb; padding: 6px 8px; text-align: left; }}
          th {{ background: #f0f0f0; }}
        </style>
      </head>
      <body>
        <p>Attached is the latest <code>results.xlsx</code> workbook.</p>
        {simulation_totals}
        <h2>{ranked_title}</h2>
        {ranked_html}
        <h2>Upcoming IPOs</h2>
        {ipo_html}
      </body>
    </html>
    """
    return subject, body


def send_email(subject: str, html_body: str) -> None:
    smtp_host = required_env("SMTP_HOST")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    smtp_username = required_env("SMTP_USERNAME")
    smtp_password = required_env("SMTP_PASSWORD")
    email_from = os.environ.get("EMAIL_FROM", smtp_username).strip() or smtp_username
    recipients = [addr.strip() for addr in required_env("EMAIL_TO").replace(";", ",").split(",") if addr.strip()]
    if not recipients:
        raise SystemExit("EMAIL_TO did not contain any recipient addresses.")

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = email_from
    message["To"] = ", ".join(recipients)
    message.set_content("Your email client does not support HTML. See the attached results workbook.")
    message.add_alternative(html_body, subtype="html")

    attach_results = os.environ.get("EMAIL_ATTACH_RESULTS", "true").strip().lower() not in {"0", "false", "no"}
    if attach_results:
        with RESULTS_PATH.open("rb") as f:
            message.add_attachment(
                f.read(),
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=RESULTS_PATH.name,
            )

    with smtplib.SMTP(smtp_host, smtp_port, timeout=60) as smtp:
        smtp.starttls()
        smtp.login(smtp_username, smtp_password)
        smtp.send_message(message)


def main() -> None:
    subject, html_body = build_email_html()
    send_email(subject, html_body)
    print("Daily screener email sent.")


if __name__ == "__main__":
    main()
