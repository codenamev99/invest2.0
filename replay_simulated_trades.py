"""
Replays every row in the "Simulation" and "AM Simulation" sheets of results.xlsx
against real Polygon 1-minute bar data, independently recomputing when the
+2%/-1% target/stop would actually have been hit, and flags rows where the
recorded exit doesn't match reality. This is meant to catch exactly the class
of bug fixed in build_investment_simulation_rows() (AM-session exit monitoring
ignoring pre-market bars) -- and, more generally, to confirm historical
simulated fills are realistic before trusting the strategy with real money.

For each row, this script:
  1. Determines the actual entry bar (parsing "Entry Time" for AM Simulation
     rows, or the regular-hours open for Simulation rows).
  2. Fetches real Polygon 1-minute bars (AM extended-hours + regular-hours, as
     appropriate) from entry date through a follow-day window.
  3. Walks those bars in order applying the same target/stop/tie-break logic
     as screen_stooq.py's intraday_exit(), and compares the result against
     what's recorded in the spreadsheet.
  4. Cross-checks the recorded entry price against the actual day's OHLC range
     (data 2/) and flags thin-looking fills (low volume relative to position
     size) as a soft liquidity warning.

This script only reads results.xlsx and data 2/; it writes nothing back to
either.

Usage:
  python3 replay_simulated_trades.py --results results.xlsx --root "data 2/daily/us"
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime, time as dtime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import requests
from openpyxl import load_workbook

import screen_stooq as ss

ET = ZoneInfo("America/New_York")
POLYGON_TICKER_URL = "https://api.polygon.io/v2/aggs/ticker/{ticker}/range/1/minute/{start}/{end}"


def resolve_path(p: str) -> Path:
    p = (p or "").strip()
    if "${workspaceFolder}" in p:
        p = p.replace("${workspaceFolder}", str(Path.cwd()))
    p = os.path.expandvars(os.path.expanduser(p))
    return Path(p).resolve()


def symbol_from_path(path: Path) -> str:
    name = path.name
    return name[:-4].upper() if name.lower().endswith(".us.txt") else path.stem.upper()


TIME_RE = re.compile(r"(\d{1,2}):(\d{2})\s*(AM|PM)", re.IGNORECASE)


def parse_entry_time(value: object) -> dtime | None:
    if not isinstance(value, str):
        return None
    m = TIME_RE.search(value)
    if not m:
        return None
    hour, minute, ampm = int(m.group(1)), int(m.group(2)), m.group(3).upper()
    if ampm == "AM":
        hour = 0 if hour == 12 else hour
    else:
        hour = 12 if hour == 12 else hour + 12
    return dtime(hour, minute)


@dataclass
class SheetRow:
    sheet: str
    symbol: str
    entry_date: date
    entry_price: float
    entry_time: dtime | None
    exit_date: date | None
    exit_time_str: str | None
    exit_price: float | None


def read_sheet_rows(wb, sheet_name: str) -> list[SheetRow]:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = [" ".join(str(h).split()) for h in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    out: list[SheetRow] = []
    for r in rows[1:]:
        symbol = r[idx["Symbol"]]
        entry_date_val = r[idx["Entry Date"]]
        entry_price = r[idx["Entry Price"]]
        if not symbol or entry_date_val is None or entry_price is None:
            continue
        entry_time = parse_entry_time(r[idx["Entry Time"]]) if "Entry Time" in idx else None
        exit_date_val = r[idx.get("Exit Date", -1)] if "Exit Date" in idx else None
        exit_time_val = r[idx.get("Exit Time", -1)] if "Exit Time" in idx else None
        exit_price_val = r[idx.get("Exit Price", -1)] if "Exit Price" in idx else None
        out.append(SheetRow(
            sheet=sheet_name,
            symbol=str(symbol).upper(),
            entry_date=entry_date_val.date() if hasattr(entry_date_val, "date") else entry_date_val,
            entry_price=float(entry_price),
            entry_time=entry_time,
            exit_date=(exit_date_val.date() if hasattr(exit_date_val, "date") else exit_date_val),
            exit_time_str=str(exit_time_val) if exit_time_val else None,
            exit_price=float(exit_price_val) if exit_price_val is not None else None,
        ))
    return out


def fetch_minute_bars(api_key: str, ticker: str, start: date, end: date, max_attempts: int = 3) -> list[dict]:
    bars: list[dict] = []
    url = POLYGON_TICKER_URL.format(ticker=ticker, start=start.isoformat(), end=end.isoformat())
    params = {"adjusted": "true", "sort": "asc", "limit": 50000, "apiKey": api_key}
    for attempt in range(1, max_attempts + 1):
        try:
            resp = requests.get(url, params=params, timeout=60)
            if resp.status_code == 429:
                time.sleep(5)
                continue
            resp.raise_for_status()
            payload = resp.json()
            break
        except requests.RequestException:
            if attempt >= max_attempts:
                return bars
            time.sleep(2)
    else:
        return bars

    for bar in payload.get("results") or []:
        ts = bar.get("t")
        if ts is None:
            continue
        ts_et = datetime.fromtimestamp(float(ts) / 1000.0, tz=ET)
        if ts_et.weekday() >= 5:
            continue
        bars.append({
            "datetime": ts_et.replace(tzinfo=None),
            "date": ts_et.date(),
            "minute_of_day": ts_et.hour * 60 + ts_et.minute,
            "open": float(bar["o"]), "high": float(bar["h"]),
            "low": float(bar["l"]), "close": float(bar["c"]),
            "volume": float(bar.get("v") or 0),
        })
    return bars


def in_am_session(m: int) -> bool:
    return 4 * 60 <= m < 9 * 60 + 30


def in_regular_session(m: int) -> bool:
    return 9 * 60 + 30 <= m < 16 * 60


def replay_row(row: SheetRow, api_key: str, follow_days: int, gain_pct: float, loss_pct: float) -> dict:
    target_price = row.entry_price * (1.0 + gain_pct)
    stop_price = row.entry_price * (1.0 - loss_pct)
    entry_session = "am" if row.sheet == "AM Simulation" and row.entry_time is not None else "regular"

    window_end = row.entry_date + timedelta(days=follow_days * 3 + 10)
    all_bars = fetch_minute_bars(api_key, row.symbol, row.entry_date, window_end)

    am_bars = [b for b in all_bars if b["date"] == row.entry_date and in_am_session(b["minute_of_day"])]
    regular_bars = [b for b in all_bars if in_regular_session(b["minute_of_day"])]
    bars = (am_bars + regular_bars) if entry_session == "am" else regular_bars

    result = {
        "symbol": row.symbol, "sheet": row.sheet, "entry_date": row.entry_date.isoformat(),
        "entry_price": row.entry_price, "recorded_exit_date": row.exit_date.isoformat() if row.exit_date else "",
        "recorded_exit_time": row.exit_time_str or "", "recorded_exit_price": row.exit_price,
        "computed_exit_date": "", "computed_exit_time": "", "computed_exit_price": "",
        "exit_match": "no_data", "entry_price_in_day_range": "unknown", "thin_fill_bar_volume": "",
    }

    if not bars:
        return result

    if row.entry_time is not None:
        entry_dt = datetime.combine(row.entry_date, row.entry_time)
    else:
        first_regular = next((b for b in regular_bars if b["date"] == row.entry_date), None)
        entry_dt = first_regular["datetime"] if first_regular else datetime.combine(row.entry_date, dtime(9, 30))

    walk_bars = [b for b in bars if b["datetime"] >= entry_dt]
    trading_days_seen = sorted({b["date"] for b in bars})

    computed_price = computed_dt = computed_reason = None
    for b in walk_bars:
        hit_target = b["high"] >= target_price
        hit_stop = b["low"] <= stop_price
        if hit_target and hit_stop:
            computed_price, computed_dt, computed_reason = stop_price, b["datetime"], "Both hit same minute - assumed -1% first"
            break
        if hit_stop:
            computed_price, computed_dt, computed_reason = stop_price, b["datetime"], "-1% stop"
            break
        if hit_target:
            computed_price, computed_dt, computed_reason = target_price, b["datetime"], "+2% target"
            break

    if computed_price is None and len(trading_days_seen) >= follow_days and walk_bars:
        last = walk_bars[-1]
        computed_price, computed_dt, computed_reason = last["close"], last["datetime"], "Max N trading days"

    if computed_dt is not None:
        result["computed_exit_date"] = computed_dt.date().isoformat()
        result["computed_exit_time"] = computed_dt.strftime("%I:%M %p ET")
        result["computed_exit_price"] = computed_price

        recorded_time = parse_entry_time(row.exit_time_str)
        date_match = row.exit_date == computed_dt.date()
        time_match = (
            recorded_time is not None
            and abs((datetime.combine(date.min, recorded_time) - datetime.combine(date.min, computed_dt.time())).total_seconds()) <= 120
        )
        price_match = row.exit_price is not None and abs(row.exit_price - computed_price) <= max(0.02, 0.001 * computed_price)
        result["exit_match"] = "match" if (date_match and time_match and price_match) else "MISMATCH"

        trigger_bar = next((b for b in walk_bars if b["datetime"] == computed_dt), None)
        if trigger_bar is not None:
            shares_needed = 10_000.0 / row.entry_price
            if trigger_bar["volume"] < shares_needed * 5:
                result["thin_fill_bar_volume"] = trigger_bar["volume"]

    return result


def entry_price_in_range(symbol_paths: dict[str, Path], row: SheetRow) -> str:
    path = symbol_paths.get(row.symbol)
    if path is None:
        return "unknown"
    d, o, h, l, c = ss.load_ohlc_from_file(path)
    date_int = int(row.entry_date.strftime("%Y%m%d"))
    matches = d == date_int
    if not matches.any():
        return "unknown"
    idx = int(matches.argmax())
    lo, hi = float(l[idx]), float(h[idx])
    buf = (hi - lo) * 0.001 + 0.005
    return "yes" if (lo - buf) <= row.entry_price <= (hi + buf) else f"NO (day range {lo}-{hi})"


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--results", default="results.xlsx")
    ap.add_argument("--root", default="${workspaceFolder}/data 2/daily/us")
    ap.add_argument("--dirs", default="nyse stocks,etfs")
    ap.add_argument("--sheets", default="Simulation,AM Simulation")
    ap.add_argument("--follow-days", type=int, default=5)
    ap.add_argument("--gain-pct", type=float, default=0.02)
    ap.add_argument("--loss-pct", type=float, default=0.01)
    ap.add_argument("--api-key", default=os.environ.get("POLYGON_API_KEY", ""))
    ap.add_argument("--sleep", type=float, default=0.25)
    ap.add_argument("--limit", type=int, default=0, help="Only replay the first N rows per sheet (0 = all).")
    ap.add_argument("--out", default="replay_report.csv")
    args = ap.parse_args()

    if not args.api_key:
        raise SystemExit("Set POLYGON_API_KEY or pass --api-key.")

    root = resolve_path(args.root)
    dirs = [d.strip() for d in args.dirs.split(",") if d.strip()]
    symbol_paths: dict[str, Path] = {}
    for d in dirs:
        sub = root / d
        if not sub.exists():
            continue
        for p in sub.glob("*.txt"):
            symbol_paths[symbol_from_path(p)] = p

    wb = load_workbook(resolve_path(args.results), data_only=True)
    all_rows: list[SheetRow] = []
    for sheet_name in [s.strip() for s in args.sheets.split(",") if s.strip()]:
        rows = read_sheet_rows(wb, sheet_name)
        if args.limit:
            rows = rows[: args.limit]
        all_rows.extend(rows)

    print(f"Replaying {len(all_rows)} row(s) against live Polygon minute data...")
    results = []
    for i, row in enumerate(all_rows, start=1):
        res = replay_row(row, args.api_key, args.follow_days, args.gain_pct, args.loss_pct)
        res["entry_price_in_day_range"] = entry_price_in_range(symbol_paths, row)
        results.append(res)
        if i % 25 == 0 or i == len(all_rows):
            print(f"[{i}/{len(all_rows)}]")
        time.sleep(args.sleep)

    out_path = resolve_path(args.out)
    fieldnames = list(results[0].keys()) if results else []
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)

    mismatches = [r for r in results if r["exit_match"] == "MISMATCH"]
    no_data = [r for r in results if r["exit_match"] == "no_data"]
    bad_entry = [r for r in results if isinstance(r["entry_price_in_day_range"], str) and r["entry_price_in_day_range"].startswith("NO")]
    thin = [r for r in results if r["thin_fill_bar_volume"]]

    print()
    print(f"Total rows replayed:        {len(results)}")
    print(f"Exit matches:               {len(results) - len(mismatches) - len(no_data)}")
    print(f"Exit MISMATCHES:            {len(mismatches)}")
    print(f"No market data to replay:   {len(no_data)}")
    print(f"Entry price outside day range: {len(bad_entry)}")
    print(f"Thin-looking fills (low volume at trigger bar): {len(thin)}")
    print()
    print(f"Full detail written to {out_path}")

    if mismatches:
        by_sheet: dict[str, int] = {}
        for r in mismatches:
            by_sheet[r["sheet"]] = by_sheet.get(r["sheet"], 0) + 1
        print("Mismatches by sheet:", by_sheet)


if __name__ == "__main__":
    main()
