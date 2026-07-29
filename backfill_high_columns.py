"""
Recompute the historical high columns in an existing results.xlsx.

Rows written before the full-file high scan landed hold a high taken from a
~1-year tail window, so they duplicate the 52-week high. This rewalks every
Daily Runs row, recomputing its high as of that row's own run date, and fills
in the Post-Gap High warning column where one applies.

Split columns are deliberately left alone: they describe current corporate
actions, not the state of the world on an old run date.

    python backfill_high_columns.py --workbook results.xlsx --root "data 2/daily/us" --dry-run
"""

from __future__ import annotations

import argparse
import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import screen_stooq as ss


HIGH_HEADERS = {"2y high", "all-time high"}
POST_GAP_HEADER = "post-gap high"
DAILY_RUNS_SHEET = "Daily Runs"


def find_columns(ws) -> tuple[int, int, int, int | None]:
    """
    Locate (symbol_col, high_col, high_days_col, post_gap_col) by header text.

    Header lookup rather than fixed offsets, because this has to run against
    workbooks written both before and after the new columns were added.
    """
    headers = [str(c.value or "").strip().lower() for c in ws[1]]

    try:
        symbol_col = headers.index("symbol") + 1
    except ValueError as exc:
        raise SystemExit("Could not find a 'Symbol' column in the Daily Runs sheet.") from exc

    high_col = next((i + 1 for i, h in enumerate(headers) if h in HIGH_HEADERS), None)
    if high_col is None:
        raise SystemExit("Could not find a '2Y High' or 'All-Time High' column.")

    post_gap_col = next((i + 1 for i, h in enumerate(headers) if h == POST_GAP_HEADER), None)
    # The days-ago cell is the unlabelled column immediately right of the high.
    return symbol_col, high_col, high_col + 1, post_gap_col


def upgrade_headers(ws, high_col: int) -> int:
    """
    Bring a pre-existing sheet up to the current column layout and return the
    Post-Gap High column index.

    Workbooks written before these columns existed end at the Earnings pair, so
    the three new columns append in the same positions the screener now writes,
    keeping later runs consistent with what this backfill produces.
    """
    if str(ws.cell(row=1, column=high_col).value or "").strip().lower() == "all-time high":
        ws.cell(row=1, column=high_col, value="2Y High")

    post_gap_col = ws.max_column + 1
    ws.cell(row=1, column=post_gap_col, value="Post-Gap High")
    ws.cell(row=2, column=post_gap_col, value=f"High since last\n>{ss.PRICE_BASIS_GAP_RATIO:g}x price gap")
    ws.cell(row=1, column=post_gap_col + 1, value="Last Split")
    ws.cell(row=2, column=post_gap_col + 1, value="Date")
    ws.cell(row=2, column=post_gap_col + 2, value="From:To")
    return post_gap_col


def coerce_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", default="results.xlsx", help="Workbook to update (default: results.xlsx)")
    ap.add_argument("--root", default="data 2/daily/us", help="Price data root")
    ap.add_argument("--out", default="", help="Write here instead of updating the workbook in place")
    ap.add_argument("--dry-run", action="store_true", help="Report changes without writing")
    args = ap.parse_args()

    workbook_path = ss.resolve_path(args.workbook)
    root = ss.resolve_path(args.root)
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path)
    if DAILY_RUNS_SHEET not in wb.sheetnames:
        raise SystemExit(f"No '{DAILY_RUNS_SHEET}' sheet in {workbook_path}")
    ws = wb[DAILY_RUNS_SHEET]

    symbol_col, high_col, high_days_col, post_gap_col = find_columns(ws)
    if post_gap_col is None:
        if args.dry_run:
            print("Note: sheet predates the new columns; a real run would add them.")
        else:
            post_gap_col = upgrade_headers(ws, high_col)
            print("Upgraded sheet to the current column layout.")

    file_map = ss.build_file_map(root)
    scan_cache: dict[tuple[str, int], Any] = {}

    changed = 0
    unchanged = 0
    skipped_no_file = 0
    skipped_no_date = 0
    flagged = 0
    examples: list[str] = []

    for row_idx in range(3, ws.max_row + 1):
        symbol = str(ws.cell(row=row_idx, column=symbol_col).value or "").strip().upper()
        run_date = coerce_date(ws.cell(row=row_idx, column=1).value)
        if not symbol:
            continue
        if run_date is None:
            skipped_no_date += 1
            continue

        path = file_map.get(f"{symbol}.US")
        if path is None:
            skipped_no_file += 1
            continue

        as_of_int = ss.date_to_int(run_date)
        cache_key = (symbol, as_of_int)
        if cache_key not in scan_cache:
            # scan_all_time_high reads this module-level flag to cut history off
            # at the row's own run date, so each row sees only what it could have.
            ss.AS_OF_DATE_INT = as_of_int
            scan_cache[cache_key] = ss.scan_all_time_high(path)
        result = scan_cache[cache_key]
        if result is None:
            skipped_no_file += 1
            continue

        old_high = ws.cell(row=row_idx, column=high_col).value
        new_high = round(result.high, 4)
        new_days = (run_date - ss.date_from_int(result.date_int)).days

        try:
            differs = old_high is None or abs(float(old_high) - new_high) > 1e-6
        except (TypeError, ValueError):
            differs = True

        if differs:
            changed += 1
            if len(examples) < 8:
                examples.append(f"  {symbol:6s} {run_date}  {old_high} -> {new_high}")
        else:
            unchanged += 1

        if result.post_gap_high is not None:
            flagged += 1

        if not args.dry_run:
            ws.cell(row=row_idx, column=high_col, value=new_high)
            ws.cell(row=row_idx, column=high_days_col, value=new_days)
            if post_gap_col is not None:
                ws.cell(
                    row=row_idx,
                    column=post_gap_col,
                    value=round(result.post_gap_high, 4) if result.post_gap_high is not None else None,
                )

    ss.AS_OF_DATE_INT = None

    print(f"Rows corrected:        {changed}")
    print(f"Rows already correct:  {unchanged}")
    print(f"Rows flagged post-gap: {flagged}")
    if skipped_no_file:
        print(f"Skipped, no price file:{skipped_no_file:>4}")
    if skipped_no_date:
        print(f"Skipped, no run date:  {skipped_no_date:>4}")
    if examples:
        print("\nSample corrections:")
        print("\n".join(examples))

    if args.dry_run:
        print("\nDry run: nothing written.")
        return

    if args.out:
        out_path = ss.resolve_path(args.out)
    else:
        out_path = workbook_path
        backup = workbook_path.with_suffix(workbook_path.suffix + ".bak")
        shutil.copy2(workbook_path, backup)
        print(f"\nBacked up original to {backup}")

    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
