from __future__ import annotations

import argparse
import calendar
import csv
import io
import os
import warnings
from bisect import bisect_left, bisect_right
from datetime import date, datetime, timedelta
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor
from pathlib import Path
from typing import Any, NamedTuple
from zoneinfo import ZoneInfo

import numpy as np
warnings.filterwarnings(
    "ignore",
    message=r"urllib3 v2 only supports OpenSSL 1\.1\.1\+.*",
)
import requests
from scipy.signal import lfilter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


AS_OF_DATE_INT: int | None = None


# -----------------------------
# Portable path handling
# -----------------------------
def resolve_path(p: str) -> Path:
    """
    Makes paths more portable by supporting:
      - ${workspaceFolder}  (Cursor/VS Code style) -> current working directory
      - ~ home expansion
      - environment variables like $HOME
      - relative paths (resolved from current working directory)
    """
    p = (p or "").strip()
    if "${workspaceFolder}" in p:
        p = p.replace("${workspaceFolder}", str(Path.cwd()))
    p = os.path.expandvars(os.path.expanduser(p))
    return Path(p).resolve()


# -----------------------------
# Fast-ish file tail reader
# -----------------------------
def read_last_lines(path: Path, n: int = 600) -> list[str]:
    """
    Read the last ~n lines of a text file efficiently by seeking from the end.
    """
    block = 64 * 1024  # 64KB
    chunks: list[bytes] = []
    target = n + 20
    newline_count = 0

    with path.open("rb") as f:
        f.seek(0, 2)  # end
        pos = f.tell()
        while pos > 0 and newline_count < target:
            read_size = min(block, pos)
            pos -= read_size
            f.seek(pos)
            chunk = f.read(read_size)
            chunks.append(chunk)
            newline_count += chunk.count(b"\n")

    # Join once at the end instead of on every loop iteration
    data = b"".join(reversed(chunks))
    lines = data.splitlines()
    tail = lines[-target:]
    return [ln.decode("utf-8", errors="ignore") for ln in tail]


def load_series_from_file(
    path: Path,
    need_rows: int,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    """
    Returns (dates_int, close, volume, high, low) sorted by date.
    date_int is YYYYMMDD (int).
    """
    lines = read_last_lines(path, n=need_rows)

    rows: list[tuple[int, float, float, float, float]] = []
    for ln in lines:
        ln = ln.strip()
        if not ln or ln.startswith("<TICKER>"):
            continue

        parts = ln.split(",")
        if len(parts) < 9:
            continue

        # TICKER, PER, DATE, TIME, OPEN, HIGH, LOW, CLOSE, VOL, ...
        if parts[1] != "D":
            continue

        try:
            date_i = int(parts[2])
            high = float(parts[5])
            low = float(parts[6])
            close = float(parts[7])
            vol = float(parts[8])
        except ValueError:
            continue

        if AS_OF_DATE_INT is not None and date_i > AS_OF_DATE_INT:
            continue

        rows.append((date_i, close, vol, high, low))

    if not rows:
        empty = np.array([], dtype=float)
        return np.array([], dtype=np.int32), empty, empty, empty, empty

    rows.sort(key=lambda x: x[0])
    d = np.array([r[0] for r in rows], dtype=np.int32)
    c = np.array([r[1] for r in rows], dtype=float)
    v = np.array([r[2] for r in rows], dtype=float)
    h = np.array([r[3] for r in rows], dtype=float)
    l = np.array([r[4] for r in rows], dtype=float)
    return d, c, v, h, l


class HistoryHigh(NamedTuple):
    """
    Highest daily high in a symbol's stored history.
    """

    high: float
    date_int: int


def scan_all_time_high(path: Path) -> HistoryHigh | None:
    """
    Scan a symbol's full file for its highest high.

    The screening path only reads a tail window (need_rows), so this must be a
    separate full-file pass or the high degrades into a ~1-year high. Ties
    resolve to the most recent date, matching max_value_and_days_ago. Returns
    None when the file has no usable daily rows.
    """
    rows: list[tuple[int, float]] = []
    with path.open("r", encoding="utf-8", errors="ignore") as f:
        for ln in f:
            ln = ln.strip()
            if not ln or ln.startswith("<TICKER>"):
                continue

            parts = ln.split(",")
            if len(parts) < 9 or parts[1] != "D":
                continue

            try:
                date_i = int(parts[2])
                high = float(parts[5])
                close = float(parts[7])
            except ValueError:
                continue

            if AS_OF_DATE_INT is not None and date_i > AS_OF_DATE_INT:
                continue
            # The close is parsed only to reject rows the screener would also
            # reject; the high is the value this scan reports.
            if not np.isfinite(high) or not np.isfinite(close):
                continue

            rows.append((date_i, high))

    if not rows:
        return None

    rows.sort(key=lambda r: r[0])

    # >= walking forward keeps the most recent date when highs tie.
    best_idx = 0
    for i in range(1, len(rows)):
        if rows[i][1] >= rows[best_idx][1]:
            best_idx = i

    return HistoryHigh(
        high=float(rows[best_idx][1]),
        date_int=int(rows[best_idx][0]),
    )


def load_ohlc_from_file(path: Path) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    """
    Returns (dates_int, open, high, low, close) sorted by date.
    Reads the full file so validation can look forward from older ranking dates.
    """
    rows: list[tuple[int, float, float, float, float]] = []
    with path.open("r", encoding="utf-8", errors="ignore") as f:
        for ln in f:
            ln = ln.strip()
            if not ln or ln.startswith("<TICKER>"):
                continue

            parts = ln.split(",")
            if len(parts) < 9 or parts[1] != "D":
                continue

            try:
                date_i = int(parts[2])
                open_ = float(parts[4])
                high = float(parts[5])
                low = float(parts[6])
                close = float(parts[7])
            except ValueError:
                continue

            rows.append((date_i, open_, high, low, close))

    if not rows:
        empty = np.array([], dtype=float)
        return np.array([], dtype=np.int32), empty, empty, empty, empty

    rows.sort(key=lambda x: x[0])
    d = np.array([r[0] for r in rows], dtype=np.int32)
    o = np.array([r[1] for r in rows], dtype=float)
    h = np.array([r[2] for r in rows], dtype=float)
    l = np.array([r[3] for r in rows], dtype=float)
    c = np.array([r[4] for r in rows], dtype=float)
    return d, o, h, l, c


# -----------------------------
# Indicators
# -----------------------------
def ema(x: np.ndarray, span: int) -> np.ndarray:
    if len(x) == 0:
        return np.empty(0, dtype=float)
    x = x.astype(float)
    alpha = 2.0 / (span + 1.0)
    # IIR filter: y[n] = alpha*x[n] + (1-alpha)*y[n-1], seeded at x[0].
    # Direct Form II transposed initial state: zi = [(1-alpha)*x[0]]
    # gives y[0] = alpha*x[0] + (1-alpha)*x[0] = x[0].
    zi = np.array([(1.0 - alpha) * x[0]])
    out, _ = lfilter([alpha], [1.0, -(1.0 - alpha)], x, zi=zi)
    return out


def rsi_wilder(close: np.ndarray, period: int = 14) -> np.ndarray:
    """
    Standard Wilder RSI.
    Returns an array the same length as close, with NaNs for early periods.
    """
    close = close.astype(float)
    n = len(close)
    if n < period + 2:
        return np.full(n, np.nan, dtype=float)

    delta = np.diff(close)
    gains = np.where(delta > 0, delta, 0.0)
    losses = np.where(delta < 0, -delta, 0.0)

    rsi = np.full(n, np.nan, dtype=float)

    # Seed: SMA of first `period` values
    avg_gain_seed = float(np.mean(gains[:period]))
    avg_loss_seed = float(np.mean(losses[:period]))

    # First RSI value corresponds to close index = period
    if avg_loss_seed == 0:
        rsi[period] = 100.0
    else:
        rsi[period] = 100.0 - 100.0 / (1.0 + avg_gain_seed / avg_loss_seed)

    # Vectorized Wilder smoothing (alpha = 1/period) for gains[period:]
    # Direct Form II transposed zi: zi = [(1-alpha)*seed] initialises the filter
    # so that the first output equals Wilder(seed, gains[period]).
    gains_tail = gains[period:]
    losses_tail = losses[period:]
    if len(gains_tail) > 0:
        alpha = 1.0 / period
        zi_g = np.array([(1.0 - alpha) * avg_gain_seed])
        zi_l = np.array([(1.0 - alpha) * avg_loss_seed])
        avg_gains, _ = lfilter([alpha], [1.0, -(1.0 - alpha)], gains_tail, zi=zi_g)
        avg_losses, _ = lfilter([alpha], [1.0, -(1.0 - alpha)], losses_tail, zi=zi_l)
        with np.errstate(divide="ignore", invalid="ignore"):
            rs = np.where(avg_losses != 0, avg_gains / avg_losses, np.inf)
            rsi[period + 1: period + 1 + len(gains_tail)] = np.where(
                avg_losses == 0, 100.0, 100.0 - 100.0 / (1.0 + rs)
            )

    return rsi


def atr_wilder(high: np.ndarray, low: np.ndarray, close: np.ndarray, period: int = 14) -> np.ndarray:
    """
    Average True Range (Wilder). Returns array aligned to close with NaNs early.
    """
    high = high.astype(float)
    low = low.astype(float)
    close = close.astype(float)
    n = len(close)
    atr = np.full(n, np.nan, dtype=float)
    if n < period + 1:
        return atr

    tr = np.empty(n, dtype=float)
    tr[0] = high[0] - low[0]
    hl = high[1:] - low[1:]
    hc = np.abs(high[1:] - close[:-1])
    lc = np.abs(low[1:] - close[:-1])
    tr[1:] = np.maximum(hl, np.maximum(hc, lc))

    # Seed: SMA of tr[1..period]
    seed = float(np.mean(tr[1: period + 1]))
    atr[period] = seed

    # Vectorized Wilder smoothing (alpha = 1/period) for tr[period+1:]
    tr_tail = tr[period + 1:]
    if len(tr_tail) > 0:
        alpha = 1.0 / period
        zi = np.array([(1.0 - alpha) * seed])
        atr_tail, _ = lfilter([alpha], [1.0, -(1.0 - alpha)], tr_tail, zi=zi)
        atr[period + 1:] = atr_tail

    return atr


def macd(close: np.ndarray, fast: int = 12, slow: int = 26, signal: int = 9) -> tuple[np.ndarray, np.ndarray]:
    close = close.astype(float)
    macd_line = ema(close, fast) - ema(close, slow)
    signal_line = ema(macd_line, signal)
    return macd_line, signal_line


def beta_from_aligned_closes(stock_close: np.ndarray, bench_close: np.ndarray) -> float:
    """
    Beta via daily returns: cov(stock, bench) / var(bench)
    Arrays must already be aligned by date.
    """
    if len(stock_close) < 35 or len(bench_close) < 35:
        return np.nan

    s = stock_close.astype(float)
    b = bench_close.astype(float)

    s_ret = np.diff(s) / s[:-1]
    b_ret = np.diff(b) / b[:-1]

    mask = np.isfinite(s_ret) & np.isfinite(b_ret)
    s_ret = s_ret[mask]
    b_ret = b_ret[mask]
    if len(b_ret) < 30:
        return np.nan

    var_b = np.var(b_ret, ddof=1)
    if var_b == 0:
        return np.nan

    cov = np.cov(s_ret, b_ret, ddof=1)[0, 1]
    return float(cov / var_b)


def month_key(date_int: int) -> int:
    """
    YYYYMM for an int date in YYYYMMDD.
    """
    return (date_int // 10000) * 100 + (date_int // 100) % 100


def date_from_int(date_int: int) -> date:
    """
    Convert YYYYMMDD int to date.
    """
    year = date_int // 10000
    month = (date_int // 100) % 100
    day = date_int % 100
    return date(year, month, day)


def date_to_int(d: date) -> int:
    """
    Convert date to YYYYMMDD int.
    """
    return d.year * 10000 + d.month * 100 + d.day


def shift_months(d: date, months: int) -> date:
    """
    Shift date by N months, clamping the day if needed.
    """
    year = d.year + (d.month - 1 + months) // 12
    month = (d.month - 1 + months) % 12 + 1
    last_day = calendar.monthrange(year, month)[1]
    day = min(d.day, last_day)
    return date(year, month, day)


def monthly_closes(dates: np.ndarray, closes: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    """
    Reduce daily series to month-end closes.
    Assumes dates are sorted ascending.
    """
    if len(dates) == 0:
        return np.array([], dtype=np.int32), np.array([], dtype=float)

    dates_int = dates.astype(np.int32)
    # YYYYMM for each date
    month_keys_arr = (dates_int // 10000) * 100 + (dates_int // 100) % 100
    # Last index for each unique month (dates are sorted, so searchsorted gives last occurrence)
    unique_months = np.unique(month_keys_arr)
    last_indices = np.searchsorted(month_keys_arr, unique_months, side="right") - 1
    return unique_months, closes[last_indices].astype(float)


# -----------------------------
# Utilities
# -----------------------------
def build_file_map(root: Path) -> dict[str, Path]:
    """
    Map 'MBSX.US' -> /path/to/mbsx.us.txt for all *.us.txt under root.
    """
    mp: dict[str, Path] = {}
    for p in root.rglob("*.us.txt"):
        sym = p.name.replace(".txt", "").upper()  # MBSX.US
        if not sym.endswith(".US"):
            sym = sym.replace(".US", "") + ".US"
        mp[sym] = p
    return mp


def find_symbol_file(root: Path, sym: str) -> Path | None:
    """
    Find a single symbol file under root (early-exit).
    """
    target = f"{sym.lower()}.txt"
    for p in root.rglob(target):
        return p
    return None


def load_tickers_csv(path: Path) -> list[str]:
    """
    Accepts:
      - CSV with header column named 'symbol' or 'ticker'
      - OR single-column CSV (no header)
    Normalizes to STQ format: 'XYZ.US'
    """
    with path.open("r", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        return []

    header = [c.strip().lower() for c in rows[0]]
    if "symbol" in header:
        col_idx = header.index("symbol")
        data_rows = rows[1:]
    elif "ticker" in header:
        col_idx = header.index("ticker")
        data_rows = rows[1:]
    else:
        col_idx = 0
        data_rows = rows

    out: set[str] = set()
    for r in data_rows:
        if not r:
            continue
        s = r[col_idx].strip().upper()
        if not s:
            continue
        if not s.endswith(".US"):
            s = f"{s}.US"
        out.add(s)

    return sorted(out)


def display_symbol(sym: str) -> str:
    """
    Output symbol without the trailing ".US" portion (case-insensitive).
    Keeps any other dots (e.g., BRK.B.US -> BRK.B).
    """
    s = sym.strip()
    if s.upper().endswith(".US"):
        return s[:-3]
    return s


def normalize_symbol(sym: str) -> str:
    s = sym.strip().upper()
    if not s:
        return s
    if not s.endswith(".US"):
        s = f"{s}.US"
    return s


def prompt_run_mode() -> tuple[str, str | None]:
    print("Run mode menu:")
    print("1) All tickers")
    print("2) Specific ticker")
    while True:
        choice = input("Select option (1/2): ").strip()
        if choice in ("1", "2"):
            break
        print("Please enter 1 or 2.")
    if choice == "2":
        while True:
            ticker = input("Enter ticker symbol (e.g., AAPL or AAPL.US): ").strip()
            if ticker:
                return "single", normalize_symbol(ticker)
            print("Ticker cannot be empty.")
    return "all", None


NASDAQ_EARNINGS_URL = "https://api.nasdaq.com/api/calendar/earnings"
NASDAQ_HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept": "application/json, text/plain, */*",
    "Referer": "https://www.nasdaq.com/",
}
ALPHAVANTAGE_IPO_URL = "https://www.alphavantage.co/query"
ALPHAVANTAGE_API_KEY = "F7HUZ9ETATI052FB"
POLYGON_AGGS_URL = "https://api.polygon.io/v2/aggs/ticker/{ticker}/range/{multiplier}/{timespan}/{from_date}/{to_date}"
POLYGON_SPLITS_URL = "https://api.polygon.io/stocks/v1/splits"
EASTERN_TZ = ZoneInfo("America/New_York")
MARKET_REGIME_SPY = "SPY.US"
MARKET_REGIME_QQQ = "QQQ.US"
MARKET_REGIME_DEFAULT_MODE = "aggressive"
MARKET_REGIME_FAST_SMA_DAYS = 20
MARKET_REGIME_SMA_DAYS = 50
MARKET_REGIME_LONG_SMA_DAYS = 200
MARKET_REGIME_MOMENTUM_DAYS = 5
MARKET_REGIME_SPY_MIN_5D_RETURN = -0.02
MARKET_REGIME_MODES = ("standard", "aggressive")
HOW_IT_WORKS_SHEET_NAME = "How It Works"
RECENT_SPLITS_SHEET_NAME = "Recent Splits (90D)"
UPCOMING_IPOS_SHEET_NAME = "Upcoming IPOs (60D)"
UPCOMING_EARNINGS_SHEET_NAME = "Upcoming Earnings (14D)"
TOP10_OHLC_SHEET_NAME = "Top 10 OHLC Tracking"
INVESTMENT_DASHBOARD_SHEET_NAME = "Investment Dashboard"
SUMMARY_SHEET_NAME = "Simulation"
AM_SIMULATION_SHEET_NAME = "AM Simulation"
PM_SIMULATION_SHEET_NAME = "PM Simulation"
SIMULATION_START_DATE = date(2026, 6, 12)
LEGACY_SUMMARY_SHEET_NAME = "Summary"
DAILY_RUNS_SHEET_NAME = "Daily Runs"
TOP10_OHLC_HIDDEN_COLUMNS = ("F", "G", "H", "M", "P")
TOP10_OHLC_TRAILING_HIDDEN_COLUMNS = ("R",)
PROTECTED_SHEET_NAMES = {
    "Single Tickers",
    HOW_IT_WORKS_SHEET_NAME,
    RECENT_SPLITS_SHEET_NAME,
    UPCOMING_IPOS_SHEET_NAME,
    UPCOMING_EARNINGS_SHEET_NAME,
    TOP10_OHLC_SHEET_NAME,
    INVESTMENT_DASHBOARD_SHEET_NAME,
    SUMMARY_SHEET_NAME,
    AM_SIMULATION_SHEET_NAME,
    PM_SIMULATION_SHEET_NAME,
    LEGACY_SUMMARY_SHEET_NAME,
    DAILY_RUNS_SHEET_NAME,
}


def _format_mmddyyyy(d: date | None) -> str:
    if not d:
        return ""
    return d.strftime("%m/%d/%Y")


def _trading_days_open(entry_date: date | None, exit_date: date | None) -> int | None:
    if not entry_date or not exit_date or exit_date < entry_date:
        return None
    return sum(
        1
        for offset in range(1, (exit_date - entry_date).days + 1)
        if (entry_date + timedelta(days=offset)).weekday() < 5
    )


def _fetch_nasdaq_day(day: date, session: requests.Session) -> tuple[date, list[dict[str, Any]]]:
    """Fetch Nasdaq earnings calendar rows for a single day (no caching)."""
    try:
        resp = session.get(
            NASDAQ_EARNINGS_URL,
            params={"date": day.isoformat()},
            headers=NASDAQ_HEADERS,
            timeout=20,
        )
        if resp.status_code != 200:
            return day, []
        payload = resp.json()
    except Exception:
        return day, []

    data = payload.get("data") if isinstance(payload, dict) else None
    rows = data.get("rows") if isinstance(data, dict) else None
    return day, (rows if isinstance(rows, list) else [])


def _nasdaq_calendar_rows(
    day: date,
    session: requests.Session,
    cache: dict[date, list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    cached = cache.get(day)
    if cached is not None:
        return cached
    _, rows = _fetch_nasdaq_day(day, session)
    cache[day] = rows
    return rows


def _extract_company_name(row: dict[str, Any]) -> str:
    name = (
        row.get("companyName")
        or row.get("company")
        or row.get("company_name")
        or row.get("name")
        or ""
    )
    return str(name).strip()


def fetch_nasdaq_earnings_dates(
    symbols: list[str],
    session: requests.Session,
    today: date | None = None,
    max_days: int = 180,
) -> dict[str, dict[str, str]]:
    today = today or date.today()
    target = {s.strip().upper() for s in symbols if s and s.strip()}
    if not target:
        return {}

    # All weekdays in [today-max_days, today+max_days]
    all_days = [
        today + timedelta(days=i)
        for i in range(-max_days, max_days + 1)
        if (today + timedelta(days=i)).weekday() < 5
    ]

    # Fetch all days in parallel (I/O-bound → threads)
    with ThreadPoolExecutor(max_workers=20) as ex:
        day_data: dict[date, list[dict[str, Any]]] = dict(
            ex.map(lambda d: _fetch_nasdaq_day(d, session), all_days)
        )

    last_dates: dict[str, date] = {}
    next_dates: dict[str, date] = {}
    company_names: dict[str, str] = {}

    for day, rows in day_data.items():
        for row in rows:
            sym = str(row.get("symbol", "")).strip().upper()
            if sym not in target:
                continue
            if sym not in company_names:
                name = _extract_company_name(row)
                if name:
                    company_names[sym] = name
            # Earliest future date → next earnings
            if day >= today:
                if sym not in next_dates or day < next_dates[sym]:
                    next_dates[sym] = day
            # Most recent past date → last earnings
            else:
                if sym not in last_dates or day > last_dates[sym]:
                    last_dates[sym] = day

    out: dict[str, dict[str, str]] = {}
    for sym in target:
        out[sym] = {
            "last": _format_mmddyyyy(last_dates.get(sym)),
            "next": _format_mmddyyyy(next_dates.get(sym)),
            "name": company_names.get(sym, ""),
        }
    return out


def fetch_polygon_splits_since(
    api_key: str,
    session: requests.Session,
    since: date,
    tracked: set[str],
) -> list[dict[str, Any]]:
    """
    Every split in the tracked universe executed on or after `since`.

    Queried without a ticker filter so a split is caught wherever it happens,
    not only in symbols that ranked that day, then narrowed to tracked symbols
    locally. Returns [] when no key is configured.
    """
    api_key = (api_key or "").strip()
    if not api_key:
        return []

    rows: list[dict[str, Any]] = []
    params: dict[str, Any] = {
        "execution_date.gte": since.isoformat(),
        "sort": "execution_date.desc",
        "limit": 1000,
        "apiKey": api_key,
    }
    url: str | None = POLYGON_SPLITS_URL

    try:
        while url:
            resp = session.get(url, params=params, timeout=30)
            resp.raise_for_status()
            payload = resp.json()
            for row in payload.get("results") or []:
                symbol = str(row.get("ticker") or "").strip().upper()
                if not symbol or (tracked and symbol not in tracked):
                    continue
                try:
                    executed = datetime.strptime(
                        str(row.get("execution_date")), "%Y-%m-%d"
                    ).date()
                except (TypeError, ValueError):
                    continue
                split_from = row.get("split_from")
                split_to = row.get("split_to")
                rows.append(
                    {
                        "symbol": symbol,
                        "execution_date": executed,
                        "ratio": (
                            ""
                            if split_from in (None, "") or split_to in (None, "")
                            else f"{_trim_split_leg(split_from)}:{_trim_split_leg(split_to)}"
                        ),
                        "adjustment_type": str(row.get("adjustment_type") or "").replace("_", " "),
                    }
                )
            url = payload.get("next_url") or None
            params = {"apiKey": api_key}
    except (requests.RequestException, ValueError) as exc:
        print(f"Warning: could not fetch recent splits from Polygon ({exc}).")

    rows.sort(key=lambda r: (r["execution_date"], r["symbol"]), reverse=True)
    return rows


def write_recent_splits_sheet(
    wb: Workbook,
    split_rows: list[dict[str, Any]],
    start_date: date,
    end_date: date,
    have_api_key: bool,
) -> None:
    """
    Create or replace a workbook tab listing splits in the tracked universe.

    Anything listed here is a symbol whose stored history predates the split and
    is therefore still on the old per-share basis until it is rescaled.
    """
    sheet_name = RECENT_SPLITS_SHEET_NAME
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for rng in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(rng))
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(title=sheet_name)

    ws.append([f"Splits in tracked symbols from {start_date.isoformat()} to {end_date.isoformat()}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color=Color(indexed=9))
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor=Color(indexed=8))

    headers = ["Symbol", "Split Date", "From:To", "Type", "Action Needed"]
    ws.append(headers)
    ws.row_dimensions[1].height = 50.85
    ws.row_dimensions[2].height = 34.85
    descriptor_fill = PatternFill(fill_type="solid", fgColor=Color(indexed=9))
    descriptor_font = Font(name="Calibri", size=11, bold=True, italic=True, color=Color(indexed=8))
    descriptor_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_red = Side(style="thin", color=Color(indexed=10))
    thick_black = Side(style="thick", color=Color(indexed=8))

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = descriptor_fill
        cell.font = descriptor_font
        cell.alignment = descriptor_align
        cell.border = Border(
            left=thin_red if col_idx == 1 else thick_black,
            right=thick_black,
        )

    if split_rows:
        for row in split_rows:
            ws.append(
                [
                    row.get("symbol", ""),
                    row.get("execution_date"),
                    row.get("ratio", ""),
                    row.get("adjustment_type", ""),
                    f"Rescale price history before {row['execution_date'].isoformat()}",
                ]
            )
    elif not have_api_key:
        ws.append(["POLYGON_API_KEY not set, so splits could not be checked.", "", "", "", ""])
    else:
        ws.append(["No splits in tracked symbols in this date range.", "", "", "", ""])

    for row_idx in range(3, ws.max_row + 1):
        ws.cell(row=row_idx, column=2).number_format = "mmm d, yyyy"

    auto_size_columns(ws, min_width=10, max_width=45)


def write_how_it_works_sheet(wb: Workbook, args: Any, avg_vol_mode: str) -> None:
    """
    Create or replace the plain-English guide tab.

    Rewritten on every run so the thresholds quoted here are the ones the run
    actually used; nothing else reads this sheet.
    """
    sheet_name = HOW_IT_WORKS_SHEET_NAME
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for rng in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(rng))
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(title=sheet_name)

    beta_window = (
        f"{args.beta_months} months" if args.beta_freq == "monthly" else f"{args.beta_lookback} days"
    )
    vol_window = (
        f"{args.avg_vol_months} months" if avg_vol_mode == "months" else f"{args.avg_vol_days} days"
    )
    regime_rule = (
        "the S&P 500 is above both its 50-day and 200-day average, its 20-day average is "
        "above its 50-day average, and it rose over the last 5 days"
        if args.market_regime_mode == "aggressive"
        else "the S&P 500 is above its 50-day average and has not fallen more than 2% "
        "over the last 5 days"
    )

    # (kind, label, text). "head" starts a section, "row" is a label/value pair,
    # "text" is a full-width paragraph, "gap" is a blank spacer.
    lines: list[tuple[str, str, str]] = [
        ("title", "How This Spreadsheet Works", ""),
        (
            "text",
            "",
            "A program runs by itself every weekday evening. It looks at every NYSE stock we "
            "track, keeps the handful that look like short-term momentum setups, ranks them, "
            "and then plays out a pretend trade on each one. No real money is involved and "
            "nothing here is a recommendation to buy or sell anything.",
        ),
        ("gap", "", ""),
        ("head", "What happens each evening", ""),
        (
            "row",
            "1. Get fresh prices",
            "Download yesterday's and today's open, high, low, close and volume for every "
            "tracked stock from Polygon.io, a market data provider.",
        ),
        (
            "row",
            "2. Screen",
            "Throw out every stock that fails any of the tests in the next section. Whatever "
            "survives is a candidate.",
        ),
        (
            "row",
            "3. Rank",
            "Give each survivor a score out of 1 and sort them best-first. The top "
            f"{args.daily_limit} go on the Daily Runs tab.",
        ),
        (
            "row",
            "4. Simulate",
            f"Pretend to buy the top {args.top_n} the next morning and follow each one until "
            "it wins, loses, or times out. Results land on the Simulation tabs.",
        ),
        (
            "row",
            "5. Email",
            "Send a summary of all of the above.",
        ),
        ("gap", "", ""),
        ("head", "Which stocks make the list", ""),
        (
            "text",
            "",
            "A stock has to pass all six tests on the same day. These are momentum tests: they "
            "look for something already moving up, traded heavily enough to get in and out of, "
            "and jumpy enough to move 2% in a few days.",
        ),
        (
            "row",
            "Enough trading",
            f"At least ${args.avg_dollar_vol_min:,.0f} of the stock changes hands on an average "
            f"day, measured over the last {vol_window}. Thin stocks are hard to sell.",
        ),
        (
            "row",
            "Moves more than the market",
            f"Beta above {args.beta_min} over {beta_window}. Beta 1.0 means it moves with the "
            "S&P 500; 1.5 means it swings about half again as hard.",
        ),
        (
            "row",
            "Rising, not overheated",
            f"RSI between {args.rsi_low} and {args.rsi_high}. RSI runs 0-100 and measures "
            "recent buying pressure. Under 50 is drifting; over 70 usually means the move is "
            "already stretched.",
        ),
        (
            "row",
            "Swings enough to matter",
            f"Average daily range above {args.atr_min_pct}% of the share price, measured over "
            f"{args.atr_period} days. A stock that moves 0.5% a day will never reach a 2% target.",
        ),
        (
            "row",
            "Trend turning up",
            "MACD above its signal line. MACD compares a "
            f"{args.macd_fast}-day and a {args.macd_slow}-day average price; when the faster "
            "one pulls ahead, the recent trend is stronger than the older one.",
        ),
        (
            "row",
            "Has real price history",
            "Enough stored days to compute all of the above.",
        ),
        ("gap", "", ""),
        ("head", "How the ranking works", ""),
        (
            "text",
            "",
            "Survivors first have to clear three extra safety checks, then get scored. A stock "
            "is skipped entirely if earnings are due within 10 days (earnings move prices for "
            "reasons this screener cannot see), if it trades under $10 million a day, if its "
            "daily swing is under 2.5%, or if its RSI is 68 or higher.",
        ),
        (
            "text",
            "",
            "Everything left gets a score from 0 to 1. Five things are graded and blended by "
            "importance:",
        ),
        (
            "row",
            "Trading volume - 30%",
            "How heavily traded it is compared with the other candidates that day. Most traded "
            "scores 1, least traded scores 0.",
        ),
        (
            "row",
            "Daily swing - 25%",
            "Full marks for a 3% to 5.5% average daily range. Calmer or wilder than that loses "
            "points. Too calm never reaches the target; too wild trips the stop loss.",
        ),
        (
            "row",
            "RSI - 20%",
            "Full marks at RSI 55, sliding to zero by 40 or 70. The middle of the range is "
            "rising without being overbought.",
        ),
        (
            "row",
            "Momentum - 15%",
            "All or nothing: full marks if the MACD trend improved over the last 5 days, zero "
            "if it did not.",
        ),
        (
            "row",
            "Repeat showings - 10%",
            "How many of the last 5 runs the stock also appeared in. Showing up 5 days straight "
            "scores 1. A setup that keeps qualifying is steadier than a one-day blip.",
        ),
        (
            "text",
            "",
            "The scores are sorted highest-first and that order becomes the Rank column. Rank 1 "
            "is the best-scoring stock of that day, not a prediction.",
        ),
        ("gap", "", ""),
        ("head", "How the pretend trades work", ""),
        (
            "row",
            "Buy",
            f"$10,000 of each of the top {args.top_n} ranked stocks. Three entry times are "
            "compared: Simulation buys at the next morning's normal 9:30am open, AM Simulation "
            "at the earlier pre-market open that same morning, and PM Simulation on the "
            "evening of the ranking day itself, in after-hours trading right after the 4pm "
            "close.",
        ),
        (
            "row",
            "Sell",
            "Whichever comes first: up 2% (a win), down 1% (a loss), or 5 trading days pass "
            "(closed at whatever it is worth). Minute-by-minute prices decide the exact moment "
            "when they are available.",
        ),
        (
            "row",
            "Market check",
            f"No new pretend buys at all unless {regime_rule}. When the whole market is falling, "
            "momentum setups tend to fail together, so those days are recorded as blocked "
            "instead of traded.",
        ),
        (
            "row",
            "Same-day both",
            "If a stock hit both +2% and -1% on the same day, the loss is assumed to have come "
            "first. That keeps the results pessimistic rather than flattering.",
        ),
        ("gap", "", ""),
        ("head", "What each tab holds", ""),
        (
            "row",
            DAILY_RUNS_SHEET_NAME,
            "One row per stock that made the cut, newest day on top, blank row between days. "
            "Row 1 names each column and row 2 gives the setting behind it.",
        ),
        (
            "row",
            SUMMARY_SHEET_NAME,
            "Every pretend trade opened at the normal market open, plus running totals of wins, "
            "losses and profit.",
        ),
        (
            "row",
            AM_SIMULATION_SHEET_NAME,
            "The same trades bought at the pre-market open instead, so the entry times can "
            "be compared.",
        ),
        (
            "row",
            PM_SIMULATION_SHEET_NAME,
            "The same trades bought the evening of the ranking day, in after-hours trading "
            "just after the 4pm close, rather than waiting for the next morning. The five-day "
            "clock still starts the following trading day.",
        ),
        (
            "row",
            UPCOMING_EARNINGS_SHEET_NAME,
            "Earnings dates in the next 14 days, flagged if the stock recently qualified. "
            "Earnings are the most common reason a setup breaks.",
        ),
        (
            "row",
            UPCOMING_IPOS_SHEET_NAME,
            "New listings expected in the next 60 days. Informational only, they have no price "
            "history to screen.",
        ),
        (
            "row",
            RECENT_SPLITS_SHEET_NAME,
            "Stock splits in the last 90 days. A split changes the share price without changing "
            "the company's value, so older stored prices for that stock are misleading until "
            "they are rescaled.",
        ),
        ("gap", "", ""),
        ("head", "Reading the Daily Runs columns", ""),
        ("row", "Symbol / Company", "Ticker and shortened company name."),
        ("row", "Rank", "Score position that day. 1 is the highest score."),
        (
            "row",
            "Times Seen",
            "How many of the previous 5 runs this stock also appeared in, from 0 to 5.",
        ),
        ("row", "Closing Price", "Last closing price, and how much it changed over 5 days."),
        (
            "row",
            "52-Week High / 2-Year High",
            "Highest close in the past year, and in all stored history (about 2 years), each "
            "with how long ago that was. Far below either means room to recover; just under "
            "means it is near the top of its range.",
        ),
        (
            "row",
            "Last 5% Higher Close",
            "The most recent day it closed at least 5% above today's price. A recent date means "
            "it has pulled back from that level; blank means today is close to the highest it "
            "has been.",
        ),
        ("row", "Beta vs S&P 500", "How hard it swings compared with the market. Above 1 is jumpier."),
        ("row", "RSI Momentum", "Buying pressure, 0-100. Roughly 50-70 is rising but not overbought."),
        ("row", "ATR Daily Swing %", "How far it typically travels in a day, as a % of its price."),
        (
            "row",
            "MACD Trend / Signal Line / MACD vs Signal",
            "Trend strength, its smoothed version, and the gap between them. MACD above signal "
            "means the recent trend is stronger than the older one; a widening gap means it is "
            "still strengthening.",
        ),
        ("row", "Avg Daily $ Volume", f"Dollars traded on an average day over the last {vol_window}."),
        ("row", "Earnings Dates", "Most recent and next scheduled earnings announcement."),
        (
            "row",
            "The % columns",
            "The narrow unlabelled column to the right of most values shows how that number "
            "changed over the last 5 days.",
        ),
        ("gap", "", ""),
        ("head", "Worth knowing", ""),
        (
            "text",
            "",
            "None of these trades are real. There is no broker connection anywhere in this "
            "project; the Simulation tabs are a record of what would have happened.",
        ),
        (
            "text",
            "",
            "Stored price history only goes back about 2 years, so anything labelled an all-time "
            "or 2-year figure only reaches as far as that.",
        ),
        (
            "text",
            "",
            "The screen is tuned for short holds of a few days. It says nothing about whether a "
            "company is worth owning - it never looks at earnings, debt, or what the business "
            "does.",
        ),
        (
            "text",
            "",
            "Every tab except this one and the Simulation history is rebuilt from scratch on "
            "each run, so edits made by hand will be overwritten.",
        ),
    ]

    black = Color(indexed=8)
    white = Color(indexed=9)
    title_font = Font(name="Calibri", size=16, bold=True, color=white)
    head_font = Font(name="Calibri", size=13, bold=True, color=white)
    label_font = Font(name="Calibri", size=11, bold=True, color=black)
    body_font = Font(name="Calibri", size=11, color=black)
    fill_black = PatternFill(fill_type="solid", fgColor=black)
    wrap_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    left_center = Alignment(horizontal="left", vertical="center")

    label_width = 30
    body_width = 100
    ws.column_dimensions["A"].width = label_width
    ws.column_dimensions["B"].width = body_width

    def wrapped_height(text: str, width: int) -> float:
        # Excel does not auto-fit wrapped rows, so estimate the line count from
        # the column width and set the height explicitly. The usable width is
        # shaved a little because a proportional font fits a variable number of
        # characters, and under-estimating the line count clips the text.
        usable = max(10, width - 3)
        lines_used = 0
        for paragraph in text.split("\n"):
            lines_used += max(1, -(-len(paragraph) // usable))
        return max(15.0, lines_used * 15.0 + 3.0)

    for kind, label, text in lines:
        row_idx = ws.max_row + 1 if ws.max_row > 1 or ws["A1"].value is not None else 1
        if kind == "gap":
            ws.cell(row=row_idx, column=1, value=None)
            ws.row_dimensions[row_idx].height = 8.0
            continue
        if kind in ("title", "head"):
            cell = ws.cell(row=row_idx, column=1, value=label)
            cell.font = title_font if kind == "title" else head_font
            cell.alignment = left_center
            for col_idx in (1, 2):
                ws.cell(row=row_idx, column=col_idx).fill = fill_black
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            ws.row_dimensions[row_idx].height = 28.0 if kind == "title" else 22.0
            continue
        if kind == "row":
            label_cell = ws.cell(row=row_idx, column=1, value=label)
            label_cell.font = label_font
            label_cell.alignment = wrap_top
            body_cell = ws.cell(row=row_idx, column=2, value=text)
            body_cell.font = body_font
            body_cell.alignment = wrap_top
            ws.row_dimensions[row_idx].height = max(
                wrapped_height(text, body_width), wrapped_height(label, label_width)
            )
            continue
        # Full-width paragraph.
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.font = body_font
        cell.alignment = wrap_top
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        ws.row_dimensions[row_idx].height = wrapped_height(text, label_width + body_width)

    ws.sheet_view.showGridLines = False


def _trim_split_leg(value: Any) -> str:
    """Render a split ratio leg without a trailing '.0' (4.0 -> '4')."""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    if number.is_integer():
        return str(int(number))
    return f"{number:g}"


def fetch_nasdaq_upcoming_earnings(
    session: requests.Session,
    start_date: date,
    end_date: date,
) -> list[dict[str, Any]]:
    """
    Fetch all Nasdaq earnings calendar rows scheduled from start_date through end_date.
    """
    if end_date < start_date:
        return []

    all_days = [
        start_date + timedelta(days=i)
        for i in range((end_date - start_date).days + 1)
    ]

    # Fetch all days in parallel
    with ThreadPoolExecutor(max_workers=20) as ex:
        day_data: dict[date, list[dict[str, Any]]] = dict(
            ex.map(lambda d: _fetch_nasdaq_day(d, session), all_days)
        )

    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, date]] = set()
    for earnings_date in all_days:
        for row in day_data.get(earnings_date, []):
            symbol = str(row.get("symbol", "")).strip().upper()
            if not symbol:
                continue
            key = (symbol, earnings_date)
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "symbol": symbol,
                    "name": _extract_company_name(row),
                    "earnings_date": earnings_date,
                    "time": str(row.get("time") or row.get("timeOfDay") or "").strip(),
                    "eps_forecast": str(row.get("epsForecast") or row.get("eps_forecast") or "").strip(),
                    "no_of_estimates": str(row.get("noOfEsts") or row.get("no_of_estimates") or "").strip(),
                    "last_year_eps": str(row.get("lastYearEPS") or row.get("last_year_eps") or "").strip(),
                }
            )

    rows.sort(key=lambda x: (x["earnings_date"], x["symbol"]))
    return rows


def fetch_alphavantage_upcoming_ipos(
    api_key: str,
    session: requests.Session,
    start_date: date,
    end_date: date,
) -> list[dict[str, Any]]:
    """
    Fetch upcoming IPOs from Alpha Vantage IPO calendar and keep only entries
    scheduled from start_date through end_date (inclusive).
    """
    if not api_key.strip():
        return []

    try:
        resp = session.get(
            ALPHAVANTAGE_IPO_URL,
            params={"function": "IPO_CALENDAR", "apikey": api_key.strip()},
            timeout=20,
        )
        if resp.status_code != 200:
            return []
        body = resp.text or ""
    except Exception:
        return []

    if not body.strip():
        return []

    rows: list[dict[str, Any]] = []
    try:
        reader = csv.DictReader(io.StringIO(body))
        for row in reader:
            ipo_text = str(row.get("ipoDate", "")).strip()
            if not ipo_text:
                continue
            try:
                ipo_dt = datetime.strptime(ipo_text, "%Y-%m-%d").date()
            except ValueError:
                continue
            if not (start_date <= ipo_dt <= end_date):
                continue

            rows.append(
                {
                    "symbol": str(row.get("symbol", "")).strip(),
                    "name": str(row.get("name", "")).strip(),
                    "ipo_date": ipo_dt,
                    "price_range": str(row.get("priceRange", "")).strip(),
                    "exchange": str(row.get("exchange", "")).strip(),
                    "currency": str(row.get("currency", "")).strip(),
                    "shares_offered": str(row.get("sharesOffered", "")).strip(),
                    "estimated_volume": str(row.get("estimatedVolume", "")).strip(),
                }
            )
    except Exception:
        return []

    rows.sort(key=lambda x: (x["ipo_date"], x["symbol"]))
    return rows


def write_upcoming_ipos_sheet(
    wb: Workbook,
    ipo_rows: list[dict[str, Any]],
    start_date: date,
    end_date: date,
) -> None:
    """
    Create or replace a workbook tab with upcoming IPOs for the selected window.
    """
    sheet_name = UPCOMING_IPOS_SHEET_NAME
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        merged_ranges = list(ws.merged_cells.ranges)
        for rng in merged_ranges:
            ws.unmerge_cells(str(rng))
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(title=sheet_name)

    ws.append([f"Upcoming IPOs from {start_date.isoformat()} to {end_date.isoformat()}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color=Color(indexed=9))
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor=Color(indexed=8))

    headers = [
        "Symbol",
        "Company",
        "IPO Date",
        "Price Range",
        "Exchange",
        "Currency",
        "Shares Offered",
        "Estimated Volume",
    ]
    ws.append(headers)
    ws.row_dimensions[1].height = 50.85
    ws.row_dimensions[2].height = 34.85
    descriptor_fill = PatternFill(fill_type="solid", fgColor=Color(indexed=9))
    descriptor_font = Font(name="Calibri", size=11, bold=True, italic=True, color=Color(indexed=8))
    descriptor_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_red = Side(style="thin", color=Color(indexed=10))
    thick_black = Side(style="thick", color=Color(indexed=8))

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = descriptor_fill
        cell.font = descriptor_font
        cell.alignment = descriptor_align
        cell.border = Border(
            left=thin_red if col_idx == 1 else thick_black,
            right=thick_black,
        )

    if ipo_rows:
        for row in ipo_rows:
            ws.append(
                [
                    row.get("symbol", ""),
                    row.get("name", ""),
                    row.get("ipo_date"),
                    row.get("price_range", ""),
                    row.get("exchange", ""),
                    row.get("currency", ""),
                    row.get("shares_offered", ""),
                    row.get("estimated_volume", ""),
                ]
            )
    else:
        ws.append(["No upcoming IPOs found in this date range.", "", "", "", "", "", "", ""])

    for row_idx in range(3, ws.max_row + 1):
        ws.cell(row=row_idx, column=3).number_format = "mmm d, yyyy"

    auto_size_columns(ws, min_width=10, max_width=45)


def write_upcoming_earnings_sheet(
    wb: Workbook,
    earnings_rows: list[dict[str, Any]],
    start_date: date,
    end_date: date,
    qualified_dates: dict[str, date],
) -> None:
    """
    Create or replace a workbook tab with upcoming earnings and current-run qualification.
    """
    sheet_name = UPCOMING_EARNINGS_SHEET_NAME
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        merged_ranges = list(ws.merged_cells.ranges)
        for rng in merged_ranges:
            ws.unmerge_cells(str(rng))
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(title=sheet_name)

    ws.append([f"Upcoming earnings from {start_date.isoformat()} to {end_date.isoformat()}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color=Color(indexed=9))
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor=Color(indexed=8))

    headers = [
        "Symbol",
        "Company",
        "Earnings Date",
        "Time",
        "EPS Forecast",
        "No. of Estimates",
        "Last Year EPS",
        "Qualified Results?",
        "Qualified Date",
    ]
    ws.append(headers)
    ws.row_dimensions[1].height = 50.85
    ws.row_dimensions[2].height = 34.85
    descriptor_fill = PatternFill(fill_type="solid", fgColor=Color(indexed=9))
    descriptor_font = Font(name="Calibri", size=11, bold=True, italic=True, color=Color(indexed=8))
    descriptor_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_red = Side(style="thin", color=Color(indexed=10))
    thick_black = Side(style="thick", color=Color(indexed=8))

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = descriptor_fill
        cell.font = descriptor_font
        cell.alignment = descriptor_align
        cell.border = Border(
            left=thin_red if col_idx == 1 else thick_black,
            right=thick_black,
        )

    if earnings_rows:
        for row in earnings_rows:
            symbol = str(row.get("symbol", "")).strip().upper()
            qualified_date = qualified_dates.get(symbol)
            ws.append(
                [
                    symbol,
                    row.get("name", ""),
                    row.get("earnings_date"),
                    row.get("time", ""),
                    row.get("eps_forecast", ""),
                    row.get("no_of_estimates", ""),
                    row.get("last_year_eps", ""),
                    "Yes" if qualified_date else "No",
                    qualified_date,
                ]
            )
    else:
        ws.append(["No upcoming earnings found in this date range.", "", "", "", "", "", "", "", ""])

    for row_idx in range(3, ws.max_row + 1):
        ws.cell(row=row_idx, column=3).number_format = "mmm d, yyyy"
        ws.cell(row=row_idx, column=9).number_format = "mmm d, yyyy"

    auto_size_columns(ws, min_width=10, max_width=45)


def fmt2(x: Any) -> str:
    """
    Format numbers to at most 2 decimal places.
    - NaN/None -> empty string
    - ints/floats -> rounded to 2 decimals, always shown with 2 decimals
    """
    if x is None:
        return ""
    try:
        if isinstance(x, (float, np.floating)) and not np.isfinite(x):
            return ""
        val = float(x)
        return f"{val:.2f}"
    except Exception:
        return str(x)


def pct_change(current: float, previous: float) -> float:
    """
    Percent change from previous to current.
    Returns NaN when previous is 0 or either value is non-finite.
    """
    if not (np.isfinite(current) and np.isfinite(previous)):
        return np.nan
    if previous == 0:
        return np.nan
    return (current - previous) / previous * 100.0


def max_value_and_days_ago(
    dates: np.ndarray,
    values: np.ndarray,
    last_date: date,
    cutoff_int: int | None = None,
) -> tuple[float | None, int | None]:
    """
    Max value (most recent if tied) and calendar days since that date.
    Optionally restrict to dates >= cutoff_int (YYYYMMDD).
    """
    if len(dates) == 0 or len(values) == 0:
        return None, None
    if cutoff_int is not None:
        mask = dates >= cutoff_int
        if not np.any(mask):
            return None, None
        dates = dates[mask]
        values = values[mask]
    max_value = float(np.max(values))
    idxs = np.where(values == max_value)[0]
    if len(idxs) == 0:
        return None, None
    max_date_int = int(dates[idxs[-1]])
    days_ago = (last_date - date_from_int(max_date_int)).days
    return max_value, int(days_ago)


def last_close_5pct_higher_info(
    dates: np.ndarray,
    closes: np.ndarray,
    last_close: float,
    last_date: date,
) -> tuple[str | None, int | None]:
    """
    Last date when close >= 5% above the most recent close.
    Returns (MM/DD/YYYY, days_ago). Excludes the most recent close.
    """
    if len(closes) < 2 or not np.isfinite(last_close):
        return "", None
    threshold = last_close * 1.05
    mask = closes[:-1] >= threshold
    if not np.any(mask):
        return "", None
    idx = int(np.where(mask)[0][-1])
    dt = date_from_int(int(dates[idx]))
    return _format_mmddyyyy(dt), int((last_date - dt).days)


def clear_sheet(ws) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    ws.delete_rows(1, ws.max_row)


def is_empty_sheet(ws) -> bool:
    """
    Returns True if the worksheet has no values.
    """
    return ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None


def append_single_ticker_section(
    ws,
    heading: str,
    header_row: list[Any],
    descriptors: list[Any],
    data_row: list[Any],
) -> None:
    if not is_empty_sheet(ws):
        ws.append([])

    ws.append([heading])
    heading_row = ws.max_row
    ws.merge_cells(start_row=heading_row, start_column=1, end_row=heading_row, end_column=len(header_row))
    heading_cell = ws.cell(row=heading_row, column=1)
    heading_cell.font = Font(name="Calibri", size=12, bold=True)
    heading_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.append([h if h is not None else "" for h in header_row])
    ws.append([d if d is not None else "" for d in descriptors])
    ws.append(data_row)

    header_row_idx = heading_row + 1
    descriptor_row_idx = heading_row + 2
    for col_idx in range(1, len(header_row) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx in range(1, len(descriptors) + 1):
        cell = ws.cell(row=descriptor_row_idx, column=col_idx)
        cell.font = Font(name="Calibri", size=10, italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _parse_run_sheet_date(sheet_name: str) -> date | None:
    base_name = sheet_name.split(" (", 1)[0].strip()
    try:
        return datetime.strptime(base_name, "%d %b %Y").date()
    except Exception:
        return None


def iter_run_rows(wb: Workbook):
    """Yield (run_date, headers, row) from consolidated and legacy run sheets."""
    if DAILY_RUNS_SHEET_NAME in wb.sheetnames:
        ws = wb[DAILY_RUNS_SHEET_NAME]
        headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
        for row in ws.iter_rows(min_row=3, values_only=True):
            run_date = _coerce_date(row[0] if row else None)
            if run_date is not None:
                yield run_date, headers[1:], row[1:]
    for name in wb.sheetnames:
        run_date = _parse_run_sheet_date(name)
        if run_date is None:
            continue
        ws = wb[name]
        headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row and row[0] is not None:
                yield run_date, headers, row


def prepare_daily_runs_sheet(wb: Workbook, header_row: list[Any]):
    """Migrate legacy date tabs into one newest-first Daily Runs table."""
    records: list[tuple[date, tuple[Any, ...]]] = []
    if DAILY_RUNS_SHEET_NAME in wb.sheetnames:
        old_ws = wb[DAILY_RUNS_SHEET_NAME]
        for row in old_ws.iter_rows(min_row=3, values_only=True):
            run_date = _coerce_date(row[0] if row else None)
            if run_date is not None:
                records.append((run_date, tuple(row[1:])))

    legacy_names: list[str] = []
    for name in wb.sheetnames:
        run_date = _parse_run_sheet_date(name)
        if run_date is None:
            continue
        legacy_names.append(name)
        old_ws = wb[name]
        for row in old_ws.iter_rows(min_row=3, values_only=True):
            if row and row[0] is not None:
                records.append((run_date, tuple(row)))

    for name in legacy_names:
        del wb[name]
    if DAILY_RUNS_SHEET_NAME in wb.sheetnames:
        ws = wb[DAILY_RUNS_SHEET_NAME]
        for merged_range in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged_range))
        clear_sheet(ws)
    elif len(wb.sheetnames) == 1 and is_empty_sheet(wb.active):
        ws = wb.active
        ws.title = DAILY_RUNS_SHEET_NAME
    else:
        ws = wb.create_sheet(DAILY_RUNS_SHEET_NAME, 0)

    ws.append(header_row)
    ws.append([None] * len(header_row))
    previous_date: date | None = None
    for run_date, row in sorted(records, key=lambda item: item[0], reverse=True):
        if previous_date is not None and run_date != previous_date:
            ws.append([None] * len(header_row))
        values = [run_date, *row]
        ws.append(values[: len(header_row)] + [None] * max(0, len(header_row) - len(values)))
        previous_date = run_date
    return ws


def count_recent_symbol_occurrences(
    wb: Workbook,
    max_runs: int = 5,
    before_date: date | None = None,
) -> dict[str, int]:
    """
    Count how many of the last N unique run dates each symbol appeared in.
    When before_date is provided, exclude sheets from that date or later so
    repeated same-day runs do not count the current run as prior history.
    """
    if max_runs <= 0 or not wb.sheetnames:
        return {}

    rows_by_date: dict[date, list[tuple[Any, ...]]] = {}
    for run_date, _headers, row in iter_run_rows(wb):
        if before_date is None or run_date < before_date:
            rows_by_date.setdefault(run_date, []).append(row)
    recent_dates = sorted(rows_by_date)[-max_runs:]
    counts: dict[str, int] = {}
    for run_date in recent_dates:
        symbols_in_sheet: set[str] = set()
        for row in rows_by_date[run_date]:
            val = row[0] if row else None
            if val is None:
                continue
            sym = str(val).strip().upper()
            if sym:
                symbols_in_sheet.add(sym)
        for sym in symbols_in_sheet:
            counts[sym] = counts.get(sym, 0) + 1

    return counts


def collect_qualified_result_dates(
    wb: Workbook,
    current_results: list[dict[str, Any]] | None = None,
    current_date: date | None = None,
) -> dict[str, date]:
    """
    Map each symbol that appeared in a regular results sheet to its latest qualifying date.
    """
    qualified_dates: dict[str, date] = {}

    for sheet_date, _headers, row in iter_run_rows(wb):
        val = row[0] if row else None
        if val is None:
            continue
        sym = str(val).strip().upper()
        if not sym:
            continue
        existing = qualified_dates.get(sym)
        if existing is None or sheet_date > existing:
            qualified_dates[sym] = sheet_date

    if current_results and current_date is not None:
        for row in current_results:
            sym = str(row.get("symbol", "")).strip().upper()
            if not sym:
                continue
            existing = qualified_dates.get(sym)
            if existing is None or current_date > existing:
                qualified_dates[sym] = current_date

    return qualified_dates


def _coerce_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d %b %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            continue
    return None


def _coerce_int(value: Any) -> int | None:
    if value is None:
        return None
    try:
        if isinstance(value, str) and not value.strip():
            return None
        return int(float(value))
    except Exception:
        return None


def _coerce_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        out = float(value)
    except Exception:
        return None
    return out if np.isfinite(out) else None


def _top10_first_threshold_hit(
    open_val: float | None,
    high_val: float | None,
    low_val: float | None,
    rank_close: float | None,
) -> str:
    if open_val is None or high_val is None or low_val is None or rank_close is None:
        return ""
    target = rank_close * 1.02
    stop = rank_close * 0.99
    if open_val >= target:
        return "+2% first"
    if open_val <= stop:
        return "-1% first"

    hit_target = high_val >= target
    hit_stop = low_val <= stop
    if hit_target and hit_stop:
        return "Both - order unknown"
    if hit_target:
        return "+2% first"
    if hit_stop:
        return "-1% first"
    return "Neither"


def collect_top_ranked_cohorts(wb: Workbook, top_n: int = 10) -> list[dict[str, Any]]:
    """
    Read workbook run sheets and return ranked ticker cohorts for OHLC tracking.
    """
    cohorts: list[dict[str, Any]] = []
    seen_cohorts: set[tuple[date, int, str]] = set()
    for rank_date, headers, row in iter_run_rows(wb):
        if rank_date < SIMULATION_START_DATE:
            continue
        try:
            rank_idx = headers.index("rank")
        except ValueError:
            continue
        # "Closing Price" is the current header; the "close $" form is what older
        # workbooks carry, and both have to resolve or the cohort is dropped and
        # the simulation sheets rebuild empty.
        close_idx = next(
            (
                idx
                for idx, header in enumerate(headers)
                if header == "closing price" or ("close" in header and "$" in header)
            ),
            None,
        )
        if close_idx is None:
            continue

        rank = _coerce_int(row[rank_idx] if rank_idx < len(row) else None)
        if rank is None or rank < 1 or rank > top_n:
            continue
        symbol = str(row[0] if row and row[0] is not None else "").strip().upper()
        if not symbol:
            continue
        cohort_key = (rank_date, rank, symbol)
        if cohort_key in seen_cohorts:
            continue
        seen_cohorts.add(cohort_key)
        cohorts.append(
            {
                "rank_date": rank_date,
                "rank": rank,
                "symbol": symbol,
                "rank_close": _coerce_float(row[close_idx] if close_idx < len(row) else None),
            }
        )

    cohorts.sort(key=lambda r: (r["rank_date"], int(r["rank"]), str(r["symbol"])))
    return cohorts


def build_top10_ohlc_tracking_rows(
    cohorts: list[dict[str, Any]],
    symbol_paths: dict[str, Path],
    root: Path,
    follow_days: int = 5,
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    ohlc_cache: dict[str, tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]] = {}

    def pct_vs_rank(value: float, rank_close: float | None) -> float | None:
        if rank_close is None or rank_close == 0:
            return None
        return float(f"{pct_change(value, rank_close):.2f}")

    for cohort in cohorts:
        symbol = str(cohort["symbol"]).strip().upper()
        normalized = normalize_symbol(symbol)
        path = symbol_paths.get(normalized)
        if path is None:
            path = find_symbol_file(root, normalized)
        if path is None:
            continue

        if normalized not in ohlc_cache:
            ohlc_cache[normalized] = load_ohlc_from_file(path)
        dates, opens, highs, lows, closes = ohlc_cache[normalized]
        if len(dates) == 0:
            continue

        rank_date = cohort["rank_date"]
        rank_int = date_to_int(rank_date)
        start_idx = int(np.searchsorted(dates, rank_int, side="right"))
        rank_close = _coerce_float(cohort.get("rank_close"))
        for offset, idx in enumerate(range(start_idx, min(start_idx + follow_days, len(dates))), start=1):
            open_val = float(opens[idx])
            high_val = float(highs[idx])
            low_val = float(lows[idx])
            close_val = float(closes[idx])
            hit_target = ""
            hit_stop = ""
            if rank_close is not None:
                hit_target = "Yes" if high_val >= rank_close * 1.02 else "No"
                hit_stop = "Yes" if low_val <= rank_close * 0.99 else "No"
            rows.append(
                [
                    rank_date,
                    int(cohort["rank"]),
                    symbol,
                    offset,
                    date_from_int(int(dates[idx])),
                    open_val,
                    high_val,
                    low_val,
                    close_val,
                    rank_close,
                    high_val,
                    pct_vs_rank(high_val, rank_close),
                    pct_vs_rank(low_val, rank_close),
                    pct_vs_rank(close_val, rank_close),
                    hit_target,
                    hit_stop,
                    _top10_first_threshold_hit(open_val, high_val, low_val, rank_close),
                ]
            )

    return rows


def write_top10_ohlc_tracking_sheet(
    wb: Workbook,
    symbol_paths: dict[str, Path],
    root: Path,
    top_n: int = 10,
) -> None:
    cohorts = collect_top_ranked_cohorts(wb, top_n=top_n)
    refresh_dates = {cohort["rank_date"] for cohort in cohorts}
    headers = [
        "Rank Date",
        "Rank",
        "Symbol",
        "Days Since Rank",
        "Date",
        "Open",
        "High",
        "Low",
        "Close",
        "Rank Close",
        "Tracking Day High",
        "High vs Initial Rank Close %",
        "Low vs Rank Close %",
        "Close vs Initial Rank Close %",
        "Hit +2%?",
        "Hit -1%?",
        "First Hit",
        "VV",
    ]

    preserved_rows: list[list[Any]] = []
    preserved_row_keys: set[tuple[Any, ...]] = set()
    if TOP10_OHLC_SHEET_NAME in wb.sheetnames:
        ws = wb[TOP10_OHLC_SHEET_NAME]
        existing_headers = [str(cell.value or "").strip() for cell in ws[1]]
        existing_header_map = {header: idx for idx, header in enumerate(existing_headers) if header}
        for row in ws.iter_rows(min_row=2, values_only=True):
            rank_date = _coerce_date(row[0] if row else None)
            if rank_date is not None and rank_date in refresh_dates:
                continue
            if any(value is not None for value in row):
                preserved_row = [
                    row[existing_header_map[header]]
                    if header in existing_header_map and existing_header_map[header] < len(row)
                    else None
                    for header in headers
                ]
                row_key = tuple(preserved_row[:5])
                if row_key in preserved_row_keys:
                    continue
                preserved_row_keys.add(row_key)
                preserved_rows.append(preserved_row)
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(title=TOP10_OHLC_SHEET_NAME)

    ws.append(headers)
    for row in preserved_rows:
        ws.append(row + [""] * (len(headers) - len(row)))
    for row in build_top10_ohlc_tracking_rows(cohorts, symbol_paths, root):
        ws.append(row)

    tracking_day_high_col = headers.index("Tracking Day High") + 1
    high_col = headers.index("High") + 1
    first_hit_col = headers.index("First Hit") + 1
    open_col = headers.index("Open") + 1
    low_col = headers.index("Low") + 1
    rank_close_col = headers.index("Rank Close") + 1
    for row_idx in range(2, ws.max_row + 1):
        tracking_day_high_cell = ws.cell(row=row_idx, column=tracking_day_high_col)
        if tracking_day_high_cell.value in (None, ""):
            tracking_day_high_cell.value = ws.cell(row=row_idx, column=high_col).value
        first_hit_cell = ws.cell(row=row_idx, column=first_hit_col)
        if first_hit_cell.value in (None, ""):
            first_hit_cell.value = _top10_first_threshold_hit(
                _coerce_float(ws.cell(row=row_idx, column=open_col).value),
                _coerce_float(ws.cell(row=row_idx, column=high_col).value),
                _coerce_float(ws.cell(row=row_idx, column=low_col).value),
                _coerce_float(ws.cell(row=row_idx, column=rank_close_col).value),
            )

    header_fill = PatternFill(fill_type="solid", fgColor=Color(indexed=8))
    header_font = Font(name="Calibri", size=11, bold=True, color=Color(indexed=9))
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(horizontal="center", vertical="center")
    hit_target_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    hit_target_vertical_border = Side(style="thin", color="A6A6A6")
    date_group_border = Side(style="thick", color="808080")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    success_label_cell = ws["S1"]
    success_label_cell.value = "Next-Day +2% Success Rate"
    success_label_cell.fill = header_fill
    success_label_cell.font = header_font
    success_label_cell.alignment = header_align

    success_rate_cell = ws["T1"]
    success_rate_cell.value = '=IFERROR(COUNTIFS($D:$D,1,$Q:$Q,"+2% first")/COUNTIFS($D:$D,1,$Q:$Q,"<>"),"")'
    success_rate_cell.fill = header_fill
    success_rate_cell.font = header_font
    success_rate_cell.alignment = header_align
    success_rate_cell.number_format = "0.00%"

    data_last_row = max(ws.max_row, 2)
    overall_success_label_cell = ws["U1"]
    overall_success_label_cell.value = "Overall +2% Success Rate"
    overall_success_label_cell.fill = header_fill
    overall_success_label_cell.font = header_font
    overall_success_label_cell.alignment = header_align

    overall_success_rate_cell = ws["V1"]
    overall_success_rate_cell.value = (
        f'=IFERROR(LET(keys,$A$2:$A${data_last_row}&"|"&$B$2:$B${data_last_row}&"|"&$C$2:$C${data_last_row},'
        f'firstHits,$Q$2:$Q${data_last_row},'
        f'days,$D$2:$D${data_last_row},'
        f'valid,firstHits<>"",'
        f'uniqueKeys,UNIQUE(FILTER(keys,valid)),'
        f'outcomes,MAP(uniqueKeys,LAMBDA(k,LET(hitMask,(keys=k)*(firstHits<>"Neither")*(firstHits<>""),'
        f'IFERROR(INDEX(SORTBY(FILTER(firstHits,hitMask),FILTER(days,hitMask),1),1),"Neither")))),'
        f'SUM(--(outcomes="+2% first"))/ROWS(uniqueKeys)),"")'
    )
    overall_success_rate_cell.fill = header_fill
    overall_success_rate_cell.font = header_font
    overall_success_rate_cell.alignment = header_align
    overall_success_rate_cell.number_format = "0.00%"

    pct_literal_format = '0.00"%"'
    date_cols = {1, 5}
    price_cols = {6, 7, 8, 9, 10, 11}
    pct_cols = {12, 13, 14}
    previous_rank_date = None
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = data_align
        first_hit = ws.cell(row=row_idx, column=17).value
        if isinstance(first_hit, str) and first_hit.strip().lower() == "+2% first":
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = hit_target_fill
                cell.border = Border(
                    left=hit_target_vertical_border,
                    right=hit_target_vertical_border,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                    diagonal=cell.border.diagonal,
                    diagonal_direction=cell.border.diagonal_direction,
                    diagonalUp=cell.border.diagonalUp,
                    diagonalDown=cell.border.diagonalDown,
                    outline=cell.border.outline,
                    vertical=cell.border.vertical,
                    horizontal=cell.border.horizontal,
                )
        for col_idx in date_cols:
            ws.cell(row=row_idx, column=col_idx).number_format = "mmm d, yyyy"
        for col_idx in price_cols:
            ws.cell(row=row_idx, column=col_idx).number_format = '"$"#,##0.00'
        for col_idx in pct_cols:
            ws.cell(row=row_idx, column=col_idx).number_format = pct_literal_format
        rank_date = ws.cell(row=row_idx, column=1).value
        if previous_rank_date is not None and rank_date != previous_rank_date:
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=date_group_border,
                    bottom=cell.border.bottom,
                    diagonal=cell.border.diagonal,
                    diagonal_direction=cell.border.diagonal_direction,
                    diagonalUp=cell.border.diagonalUp,
                    diagonalDown=cell.border.diagonalDown,
                    outline=cell.border.outline,
                    vertical=cell.border.vertical,
                    horizontal=cell.border.horizontal,
                )
        previous_rank_date = rank_date

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = header_align if cell.row == 1 else data_align

    ws.freeze_panes = "A2"
    auto_size_columns(ws, min_width=9, max_width=22)
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].hidden = False
    # Set to an empty tuple to show all OHLC tracking columns again.
    for col_letter in TOP10_OHLC_HIDDEN_COLUMNS:
        ws.column_dimensions[col_letter].hidden = True
    # Excel always has additional blank columns; keep the first trailing one out of view.
    for col_letter in TOP10_OHLC_TRAILING_HIDDEN_COLUMNS:
        ws.column_dimensions[col_letter].hidden = True


def _market_symbol_series(
    symbol: str,
    symbol_paths: dict[str, Path],
    root: Path,
    cache: dict[str, tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]],
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray] | None:
    normalized = normalize_symbol(symbol)
    if normalized not in cache:
        path = symbol_paths.get(normalized) or find_symbol_file(root, normalized)
        if path is None:
            return None
        cache[normalized] = load_ohlc_from_file(path)
    return cache[normalized]


def evaluate_market_regime(
    rank_date: date,
    symbol_paths: dict[str, Path],
    root: Path,
    cache: dict[str, tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]],
    mode: str = MARKET_REGIME_DEFAULT_MODE,
) -> dict[str, Any]:
    """
    SPY-only market gate for new entries. Standard mode requires SPY above
    its 50DMA with 5D return > -2%; aggressive adds long-term trend checks
    and requires positive 5D momentum.
    """
    mode = mode if mode in MARKET_REGIME_MODES else MARKET_REGIME_DEFAULT_MODE
    fast_sma_days = MARKET_REGIME_FAST_SMA_DAYS
    sma_days = MARKET_REGIME_SMA_DAYS
    long_sma_days = MARKET_REGIME_LONG_SMA_DAYS
    momentum_days = MARKET_REGIME_MOMENTUM_DAYS
    min_spy_return_5d = 0.0 if mode == "aggressive" else MARKET_REGIME_SPY_MIN_5D_RETURN
    metrics: dict[str, Any] = {
        "regime": "Unknown",
        "market_regime_mode": mode,
        "entry_allowed": True,
        "reason": "Market data unavailable; entry not blocked.",
        "regime_date": rank_date,
        "spy_close": None,
        "spy_sma20": None,
        "spy_sma50": None,
        "spy_sma200": None,
        "qqq_close": None,
        "qqq_sma20": None,
        "qqq_sma50": None,
        "spy_return_5d": None,
        "qqq_return_5d": None,
    }
    rank_date_int = date_to_int(rank_date)
    spy_series = _market_symbol_series(MARKET_REGIME_SPY, symbol_paths, root, cache)
    qqq_series = _market_symbol_series(MARKET_REGIME_QQQ, symbol_paths, root, cache)
    if spy_series is None:
        metrics["reason"] = f"Missing market data for {MARKET_REGIME_SPY}; entry not blocked."
        return metrics

    spy_dates, _spy_opens, _spy_highs, _spy_lows, spy_closes = spy_series
    spy_idx = int(np.searchsorted(spy_dates, rank_date_int, side="right") - 1)
    required_spy_idx = max(sma_days - 1, momentum_days)
    if mode == "aggressive":
        required_spy_idx = max(required_spy_idx, long_sma_days - 1)
    if spy_idx < required_spy_idx:
        metrics["reason"] = "Insufficient SPY history; entry not blocked."
        return metrics

    spy_close = float(spy_closes[spy_idx])
    spy_sma20 = float(np.mean(spy_closes[spy_idx - (fast_sma_days - 1) : spy_idx + 1]))
    spy_sma50 = float(np.mean(spy_closes[spy_idx - (sma_days - 1) : spy_idx + 1]))
    spy_sma200 = (
        float(np.mean(spy_closes[spy_idx - (long_sma_days - 1) : spy_idx + 1]))
        if spy_idx >= long_sma_days - 1
        else None
    )
    spy_prev_close = float(spy_closes[spy_idx - momentum_days])
    spy_return_5d = (spy_close / spy_prev_close) - 1.0 if spy_prev_close > 0 else np.nan

    qqq_close = qqq_sma20 = qqq_sma50 = qqq_return_5d = None
    if qqq_series is not None:
        qqq_dates, _qqq_opens, _qqq_highs, _qqq_lows, qqq_closes = qqq_series
        qqq_idx = int(np.searchsorted(qqq_dates, rank_date_int, side="right") - 1)
        if qqq_idx >= sma_days - 1:
            qqq_close = float(qqq_closes[qqq_idx])
            qqq_sma20 = float(np.mean(qqq_closes[qqq_idx - (fast_sma_days - 1) : qqq_idx + 1]))
            qqq_sma50 = float(np.mean(qqq_closes[qqq_idx - (sma_days - 1) : qqq_idx + 1]))
            qqq_prev_close = float(qqq_closes[qqq_idx - momentum_days])
            qqq_return_5d = (qqq_close / qqq_prev_close) - 1.0 if qqq_prev_close > 0 else np.nan

    spy_above_sma = spy_close > spy_sma50
    spy_momentum_ok = np.isfinite(spy_return_5d) and spy_return_5d > min_spy_return_5d
    spy_above_long_sma = spy_sma200 is not None and spy_close > spy_sma200
    spy_ma_aligned = spy_sma20 > spy_sma50
    entry_allowed = spy_above_sma and spy_momentum_ok
    if mode == "aggressive":
        entry_allowed = (
            entry_allowed
            and spy_above_long_sma
            and spy_ma_aligned
        )
    failed_checks = []
    if not spy_above_sma:
        failed_checks.append("SPY <= 50DMA")
    if not spy_momentum_ok:
        failed_checks.append("SPY 5D <= 0%" if mode == "aggressive" else "SPY 5D <= -2%")
    if mode == "aggressive":
        if not spy_above_long_sma:
            failed_checks.append("SPY <= 200DMA")
        if not spy_ma_aligned:
            failed_checks.append("SPY 20DMA <= 50DMA")
    regime = "Risk-On"
    if not entry_allowed:
        regime = "Risk-Off" if len(failed_checks) >= 2 or not spy_momentum_ok else "Neutral"

    metrics.update(
        {
            "regime": regime,
            "market_regime_mode": mode,
            "entry_allowed": entry_allowed,
            "reason": "All market checks passed." if entry_allowed else "; ".join(failed_checks),
            "regime_date": date_from_int(int(spy_dates[spy_idx])),
            "spy_close": spy_close,
            "spy_sma20": spy_sma20,
            "spy_sma50": spy_sma50,
            "spy_sma200": spy_sma200,
            "qqq_close": qqq_close,
            "qqq_sma20": qqq_sma20,
            "qqq_sma50": qqq_sma50,
            "spy_return_5d": spy_return_5d,
            "qqq_return_5d": qqq_return_5d,
        }
    )
    return metrics


def build_investment_simulation_rows(
    cohorts: list[dict[str, Any]],
    symbol_paths: dict[str, Path],
    root: Path,
    position_size: float = 10_000.0,
    follow_days: int = 5,
    gain_pct: float = 0.02,
    loss_pct: float = 0.01,
    polygon_api_key: str = "",
    intraday_exit_source: str = "auto",
    market_regimes: dict[date, dict[str, Any]] | None = None,
    entry_session: str = "regular",
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    ohlc_cache: dict[str, tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]] = {}
    intraday_cache: dict[tuple[str, date, date, str], list[dict[str, Any]]] = {}
    market_series = _market_symbol_series(MARKET_REGIME_SPY, symbol_paths, root, ohlc_cache)
    market_dates = market_series[0] if market_series is not None else np.array([], dtype=np.int32)

    def polygon_ticker(symbol: str) -> str:
        return normalize_symbol(symbol).removesuffix(".US")

    def fetch_polygon_minute_bars(
        symbol: str,
        start_date: date,
        end_date: date,
        session: str = "regular",
    ) -> list[dict[str, Any]]:
        key = (normalize_symbol(symbol), start_date, end_date, session)
        if key in intraday_cache:
            return intraday_cache[key]

        bars: list[dict[str, Any]] = []
        if not polygon_api_key:
            intraday_cache[key] = bars
            return bars

        url = POLYGON_AGGS_URL.format(
            ticker=polygon_ticker(symbol),
            multiplier=1,
            timespan="minute",
            from_date=start_date.isoformat(),
            to_date=end_date.isoformat(),
        )
        params: dict[str, Any] = {
            "adjusted": "true",
            "sort": "asc",
            "limit": 50000,
            "apiKey": polygon_api_key,
        }
        while url:
            try:
                resp = requests.get(url, params=params, timeout=60)
                if resp.status_code in {401, 403}:
                    break
                resp.raise_for_status()
                payload = resp.json()
            except Exception:
                bars = []
                break

            for bar in payload.get("results") or []:
                ts = bar.get("t")
                if ts is None:
                    continue
                ts_et = datetime.fromtimestamp(float(ts) / 1000.0, tz=EASTERN_TZ)
                if ts_et.weekday() >= 5:
                    continue
                # Polygon timestamps identify the start of each minute. AM extended
                # hours are 4:00-9:29 a.m. ET; regular hours are 9:30 a.m.-4:00 p.m. ET;
                # PM extended hours are 4:00-8:00 p.m. ET.
                minute_of_day = ts_et.hour * 60 + ts_et.minute
                if session == "am":
                    in_requested_session = 4 * 60 <= minute_of_day < 9 * 60 + 30
                elif session == "pm":
                    in_requested_session = 16 * 60 <= minute_of_day < 20 * 60
                else:
                    in_requested_session = 9 * 60 + 30 <= minute_of_day < 16 * 60
                if not in_requested_session:
                    continue
                try:
                    bars.append(
                        {
                            "datetime": ts_et.replace(tzinfo=None),
                            "date": ts_et.date(),
                            "open": float(bar["o"]),
                            "high": float(bar["h"]),
                            "low": float(bar["l"]),
                            "close": float(bar["c"]),
                        }
                    )
                except (KeyError, TypeError, ValueError):
                    continue

            next_url = payload.get("next_url")
            url = str(next_url) if next_url else ""
            params = {"apiKey": polygon_api_key} if url else {}

        intraday_cache[key] = bars
        return bars

    def intraday_exit(
        symbol: str,
        entry_date: date,
        daily_dates: np.ndarray,
        start_idx: int,
        target_price: float,
        stop_price: float,
        entry_dt: datetime | None = None,
    ) -> tuple[float | None, datetime | None, str, str]:
        if intraday_exit_source == "daily" or not polygon_api_key:
            return None, None, "", ""

        end_idx = min(start_idx + follow_days, len(daily_dates))
        if end_idx <= start_idx:
            return None, None, "", ""
        end_date = date_from_int(int(daily_dates[end_idx - 1]))

        if entry_session == "pm":
            # A PM entry fills in the rank date's after-hours session, so start_idx
            # already points at the next trading day. That evening's bars from the
            # fill minute onward are scanned first, and the rank date's regular
            # session -- which ran entirely before the fill -- is never fetched.
            scan_start_date = date_from_int(int(daily_dates[start_idx]))
            pm_bars = [
                bar
                for bar in fetch_polygon_minute_bars(symbol, entry_date, entry_date, session="pm")
                if entry_dt is None or bar["datetime"] >= entry_dt
            ]
            bars = pm_bars + fetch_polygon_minute_bars(symbol, scan_start_date, end_date)
            # The rank date's evening is part of the entry day, not one of the five
            # follow days, so it must not count toward the timeout.
            follow_dates = {bar["date"] for bar in bars if bar["date"] > entry_date}
        else:
            bars = fetch_polygon_minute_bars(symbol, entry_date, end_date)
            if entry_session == "am":
                # An AM entry can fill hours before the regular-hours bars fetched
                # above begin; without the entry day's AM bars too, a target/stop
                # crossing between the AM fill and the 9:30 open is never seen.
                am_bars = fetch_polygon_minute_bars(symbol, entry_date, entry_date, session="am")
                bars = am_bars + bars
            follow_dates = {bar["date"] for bar in bars}
        if not bars:
            return None, None, "", ""

        if entry_session == "pm":
            first_bar = bars[0]
        else:
            first_bar = next((bar for bar in bars if bar["date"] >= entry_date), None)
        if first_bar is None:
            return None, None, "", ""

        for bar in bars:
            bar_dt = bar["datetime"]
            if bar_dt < first_bar["datetime"]:
                continue
            hit_target = float(bar["high"]) >= target_price
            hit_stop = float(bar["low"]) <= stop_price
            if hit_target and hit_stop:
                return stop_price, bar_dt, "Both hit same minute - assumed -1% first", "Polygon 1-min"
            if hit_stop:
                return stop_price, bar_dt, "-1% stop", "Polygon 1-min"
            if hit_target:
                return target_price, bar_dt, "+2% target", "Polygon 1-min"

        if len(follow_dates) >= follow_days:
            final_bar = bars[-1]
            return float(final_bar["close"]), final_bar["datetime"], "Max 5 trading days", "Polygon 1-min"

        return None, None, "", ""

    def append_ignored_row(
        cohort: dict[str, Any],
        symbol: str,
        entry_date: date | None,
        reason: str,
        market_regime: dict[str, Any],
    ) -> None:
        rows.append(
            {
                "rank_date": cohort["rank_date"],
                "rank": int(cohort["rank"]),
                "symbol": symbol,
                "entry_date": entry_date,
                "entry_time": "",
                "entry_price": None,
                "shares": 0.0,
                "investment": 0.0,
                "exit_date": None,
                "exit_time": "",
                "exit_price": None,
                "exit_reason": reason,
                "status": "Ignored",
                "result_currency": None,
                "result_pct": None,
                "data_source": "Unavailable",
                "market_regime": market_regime.get("regime", ""),
                "market_entry_allowed": "No",
                "market_reason": reason,
            }
        )

    for cohort in cohorts:
        symbol = str(cohort["symbol"]).strip().upper()
        normalized = normalize_symbol(symbol)
        rank_date = cohort["rank_date"]
        market_regime = (market_regimes or {}).get(rank_date, {})
        if entry_session == "pm":
            # A PM entry fills in the rank date's own after-hours session, so unlike
            # the other sessions it never waits for the next trading day.
            required_entry_date_int = date_to_int(rank_date)
            required_entry_date = rank_date
        else:
            next_market_idx = int(np.searchsorted(market_dates, date_to_int(rank_date), side="right"))
            if next_market_idx >= len(market_dates):
                append_ignored_row(
                    cohort,
                    symbol,
                    None,
                    "Ignored: next trading day is not available yet.",
                    market_regime,
                )
                continue
            required_entry_date_int = int(market_dates[next_market_idx])
            required_entry_date = date_from_int(required_entry_date_int)

        path = symbol_paths.get(normalized)
        if path is None:
            path = find_symbol_file(root, normalized)
        if path is None:
            append_ignored_row(
                cohort,
                symbol,
                required_entry_date,
                "Ignored: no local daily price file is available for this symbol.",
                market_regime,
            )
            continue

        if normalized not in ohlc_cache:
            ohlc_cache[normalized] = load_ohlc_from_file(path)
        dates, opens, highs, lows, closes = ohlc_cache[normalized]
        if len(dates) == 0:
            append_ignored_row(
                cohort,
                symbol,
                required_entry_date,
                "Ignored: the local price file contains no usable daily data.",
                market_regime,
            )
            continue

        entry_allowed = bool(market_regime.get("entry_allowed", True))
        start_idx = int(
            np.searchsorted(
                dates,
                date_to_int(rank_date),
                side="left" if entry_session == "pm" else "right",
            )
        )
        if start_idx >= len(dates) or int(dates[start_idx]) != required_entry_date_int:
            append_ignored_row(
                cohort,
                symbol,
                required_entry_date,
                f"Ignored: daily OHLC is missing for the required {required_entry_date:%b %d, %Y} entry.",
                market_regime,
            )
            continue

        entry_date = date_from_int(int(dates[start_idx]))
        entry_price = float(opens[start_idx])
        entry_time = "Market Open"
        data_source = "Daily OHLC"
        entry_fallback_reason = ""
        entry_dt: datetime | None = None
        if entry_session == "pm":
            pm_bars = (
                fetch_polygon_minute_bars(symbol, entry_date, entry_date, session="pm")
                if polygon_api_key
                else []
            )
            first_pm_bar = next((bar for bar in pm_bars if bar["date"] == entry_date), None)
            if first_pm_bar is None:
                entry_price = float(closes[start_idx])
                entry_time = "Market Close"
                data_source = "Daily close PM fallback"
                entry_fallback_reason = (
                    "Good - PM extended-hours data unavailable; used the rank date's "
                    "closing price."
                )
            else:
                entry_price = float(first_pm_bar["open"])
                entry_dt = first_pm_bar["datetime"]
                entry_time = first_pm_bar["datetime"].strftime("%I:%M %p ET")
                data_source = "Polygon 1-min PM extended hours"
        elif entry_session == "am":
            minute_bars = (
                fetch_polygon_minute_bars(symbol, entry_date, entry_date, session="am")
                if polygon_api_key
                else []
            )
            first_am_bar = next((bar for bar in minute_bars if bar["date"] == entry_date), None)
            if first_am_bar is None:
                regular_bars = (
                    fetch_polygon_minute_bars(symbol, entry_date, entry_date, session="regular")
                    if polygon_api_key
                    else []
                )
                first_regular_bar = next(
                    (bar for bar in regular_bars if bar["date"] == entry_date),
                    None,
                )
                if first_regular_bar is not None:
                    entry_price = float(first_regular_bar["open"])
                    entry_time = first_regular_bar["datetime"].strftime("%I:%M %p ET")
                    data_source = "Polygon 1-min regular-hours fallback"
                else:
                    entry_price = float(opens[start_idx])
                    entry_time = "Market Open"
                    data_source = "Daily OHLC regular-hours fallback"
                entry_fallback_reason = (
                    "Good - AM extended-hours data unavailable; used the first available "
                    "regular-hours price."
                )
            else:
                entry_price = float(first_am_bar["open"])
                entry_time = first_am_bar["datetime"].strftime("%I:%M %p ET")
                data_source = "Polygon 1-min AM extended hours"
        elif intraday_exit_source != "daily" and polygon_api_key:
            minute_bars = fetch_polygon_minute_bars(symbol, entry_date, entry_date)
            first_regular_bar = next((bar for bar in minute_bars if bar["date"] == entry_date), None)
            if first_regular_bar is not None:
                entry_price = float(first_regular_bar["open"])
                entry_time = first_regular_bar["datetime"].strftime("%I:%M %p ET")
                data_source = "Polygon 1-min"
        if entry_price <= 0:
            append_ignored_row(
                cohort,
                symbol,
                entry_date,
                "Ignored: the required entry price is missing or invalid.",
                market_regime,
            )
            continue

        investment_amount = position_size if entry_allowed else 0.0
        hypothetical_shares = position_size / entry_price
        shares = hypothetical_shares if entry_allowed else 0.0
        target_price = entry_price * (1.0 + gain_pct)
        stop_price = entry_price * (1.0 - loss_pct)
        # A PM fill lands after the rank date's regular session, so that day's own
        # high and low were set before the entry existed and would fabricate exits.
        # Daily-OHLC scanning therefore starts the following trading day.
        exit_scan_start_idx = start_idx + 1 if entry_session == "pm" else start_idx
        end_idx = min(exit_scan_start_idx + follow_days, len(dates))

        status = "Open" if entry_allowed else "Blocked"
        exit_date: date | None = None
        exit_time = ""
        exit_price: float | None = None
        exit_reason = "Open - waiting for threshold or day 5" if entry_allowed else "Market regime blocked entry"

        intraday_price, intraday_dt, intraday_reason, intraday_source = intraday_exit(
            symbol,
            entry_date,
            dates,
            exit_scan_start_idx,
            target_price,
            stop_price,
            entry_dt=entry_dt,
        )
        if intraday_price is not None and intraday_dt is not None:
            if entry_allowed:
                status = "Closed"
            exit_date = intraday_dt.date()
            exit_time = intraday_dt.strftime("%I:%M %p ET")
            exit_price = intraday_price
            exit_reason = intraday_reason
            data_source = intraday_source
        else:
            for idx in range(exit_scan_start_idx, end_idx):
                high_val = float(highs[idx])
                low_val = float(lows[idx])
                hit_target = high_val >= target_price
                hit_stop = low_val <= stop_price
                if hit_target and hit_stop:
                    if entry_allowed:
                        status = "Closed"
                    exit_date = date_from_int(int(dates[idx]))
                    exit_time = "Unavailable with daily OHLC"
                    exit_price = stop_price
                    exit_reason = "Both hit same day - assumed -1% first"
                    break
                if hit_stop:
                    if entry_allowed:
                        status = "Closed"
                    exit_date = date_from_int(int(dates[idx]))
                    exit_time = "Unavailable with daily OHLC"
                    exit_price = stop_price
                    exit_reason = "-1% stop"
                    break
                if hit_target:
                    if entry_allowed:
                        status = "Closed"
                    exit_date = date_from_int(int(dates[idx]))
                    exit_time = "Unavailable with daily OHLC"
                    exit_price = target_price
                    exit_reason = "+2% target"
                    break

        if exit_price is None and end_idx - exit_scan_start_idx >= follow_days:
            final_idx = end_idx - 1
            if entry_allowed:
                status = "Closed"
            exit_date = date_from_int(int(dates[final_idx]))
            exit_time = "Market Close"
            exit_price = float(closes[final_idx])
            exit_reason = "Max 5 trading days"

        result_currency = (hypothetical_shares * exit_price) - position_size if exit_price is not None else None
        result_pct = (exit_price / entry_price) - 1.0 if exit_price is not None else None
        rows.append(
            {
                "rank_date": rank_date,
                "rank": int(cohort["rank"]),
                "symbol": symbol,
                "entry_date": entry_date,
                "entry_time": entry_time,
                "entry_price": entry_price,
                "shares": shares,
                "investment": investment_amount,
                "exit_date": exit_date,
                "exit_time": exit_time,
                "exit_price": exit_price,
                "exit_reason": exit_reason,
                "status": status,
                "result_currency": result_currency,
                "result_pct": result_pct,
                "data_source": data_source,
                "market_regime": market_regime.get("regime", ""),
                "market_entry_allowed": "Yes" if entry_allowed else "No",
                "market_reason": market_regime.get("reason", ""),
                "entry_fallback_reason": entry_fallback_reason,
            }
        )

    rows.sort(key=lambda r: (r["rank_date"], int(r["rank"]), str(r["symbol"])))
    return rows


def write_investment_dashboard_sheet(
    wb: Workbook,
    symbol_paths: dict[str, Path],
    root: Path,
    top_n: int = 10,
    portfolio_capital: float = 500_000.0,
    position_size: float = 10_000.0,
    polygon_api_key: str = "",
    intraday_exit_source: str = "auto",
    market_regime_mode: str = MARKET_REGIME_DEFAULT_MODE,
) -> None:
    cohorts = collect_top_ranked_cohorts(wb, top_n=top_n)
    market_cache: dict[str, tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]] = {}
    market_regimes = {
        rank_date: evaluate_market_regime(rank_date, symbol_paths, root, market_cache, mode=market_regime_mode)
        for rank_date in {cohort["rank_date"] for cohort in cohorts}
    }
    simulation_rows = build_investment_simulation_rows(
        cohorts,
        symbol_paths,
        root,
        position_size=position_size,
        polygon_api_key=polygon_api_key,
        intraday_exit_source=intraday_exit_source,
        market_regimes=market_regimes,
    )

    if INVESTMENT_DASHBOARD_SHEET_NAME in wb.sheetnames:
        ws = wb[INVESTMENT_DASHBOARD_SHEET_NAME]
        for merged_range in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged_range))
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(title=INVESTMENT_DASHBOARD_SHEET_NAME)

    closed_rows = [row for row in simulation_rows if row["status"] == "Closed"]
    open_rows = [row for row in simulation_rows if row["status"] == "Open"]
    blocked_rows = [row for row in simulation_rows if row["status"] == "Blocked"]
    realized_pl = sum(float(row["result_currency"] or 0.0) for row in closed_rows)
    open_investment = sum(float(row["investment"]) for row in open_rows)
    portfolio_value = portfolio_capital + realized_pl
    portfolio_return = (portfolio_value / portfolio_capital) - 1.0 if portfolio_capital else None

    active_counts: list[int] = []
    check_dates = sorted(
        {
            d
            for row in simulation_rows
            for d in (row["entry_date"], row["exit_date"])
            if isinstance(d, date)
        }
    )
    for check_date in check_dates:
        active = 0
        for row in simulation_rows:
            if row["status"] == "Blocked":
                continue
            entry_date = row["entry_date"]
            exit_date = row["exit_date"] if isinstance(row["exit_date"], date) else check_dates[-1]
            if entry_date <= check_date <= exit_date:
                active += 1
        active_counts.append(active)
    max_concurrent_positions = max(active_counts) if active_counts else 0
    max_capital_deployed = max_concurrent_positions * position_size
    current_market_regime = (
        market_regimes[max(market_regimes)]
        if market_regimes
        else {
            "regime": "Unknown",
            "market_regime_mode": market_regime_mode,
            "entry_allowed": True,
            "reason": "No ranked cohorts found.",
            "regime_date": None,
            "spy_close": None,
            "spy_sma20": None,
            "spy_sma50": None,
            "spy_sma200": None,
            "qqq_close": None,
            "qqq_sma20": None,
            "qqq_sma50": None,
            "spy_return_5d": None,
            "qqq_return_5d": None,
        }
    )

    summary_rows = [
        ("Initial Portfolio Capital", portfolio_capital),
        ("Position Size Per Entry", position_size),
        ("Total Entries", len(simulation_rows) - len(blocked_rows)),
        ("Closed Exits", len(closed_rows)),
        ("Open Positions", len(open_rows)),
        ("Blocked Entries", len(blocked_rows)),
        ("Open Position Cost", open_investment),
        ("Realized P/L", realized_pl),
        ("Portfolio Value (Realized)", portfolio_value),
        ("Portfolio Return (Realized)", portfolio_return),
        ("Max Concurrent Positions", max_concurrent_positions),
        ("Max Capital Deployed", max_capital_deployed),
        ("Current Market Regime", current_market_regime.get("regime")),
        ("Market Regime Mode", current_market_regime.get("market_regime_mode")),
        ("Market Regime Date", current_market_regime.get("regime_date")),
        ("New Entries Allowed", "Yes" if current_market_regime.get("entry_allowed", True) else "No"),
        ("Market Regime Reason", current_market_regime.get("reason")),
        (
            "SPY Close / 50DMA",
            (
                f"{current_market_regime['spy_close']:.2f} / {current_market_regime['spy_sma50']:.2f}"
                if current_market_regime.get("spy_close") is not None
                and current_market_regime.get("spy_sma50") is not None
                else None
            ),
        ),
        (
            "SPY 20DMA / 50DMA",
            (
                f"{current_market_regime['spy_sma20']:.2f} / {current_market_regime['spy_sma50']:.2f}"
                if current_market_regime.get("spy_sma20") is not None
                and current_market_regime.get("spy_sma50") is not None
                else None
            ),
        ),
        (
            "SPY Close / 200DMA",
            (
                f"{current_market_regime['spy_close']:.2f} / {current_market_regime['spy_sma200']:.2f}"
                if current_market_regime.get("spy_close") is not None
                and current_market_regime.get("spy_sma200") is not None
                else None
            ),
        ),
        (
            "QQQ Close / 50DMA",
            (
                f"{current_market_regime['qqq_close']:.2f} / {current_market_regime['qqq_sma50']:.2f}"
                if current_market_regime.get("qqq_close") is not None
                and current_market_regime.get("qqq_sma50") is not None
                else None
            ),
        ),
        (
            "QQQ 20DMA / 50DMA",
            (
                f"{current_market_regime['qqq_sma20']:.2f} / {current_market_regime['qqq_sma50']:.2f}"
                if current_market_regime.get("qqq_sma20") is not None
                and current_market_regime.get("qqq_sma50") is not None
                else None
            ),
        ),
        ("SPY 5D Return", current_market_regime.get("spy_return_5d")),
        ("QQQ 5D Return", current_market_regime.get("qqq_return_5d")),
    ]
    summary_label_rows = {label: row_idx for row_idx, (label, _value) in enumerate(summary_rows, start=2)}

    ws.append(["Overall Portfolio Standings", ""])
    for label, value in summary_rows:
        ws.append([label, value])
    ws.append([])
    ws.append(
        [
            "Note",
            "Entry is next trading day's regular-market open after the rank date only when the market regime permits new entries. With POLYGON_API_KEY, Exit Time is the first Polygon 1-minute regular-session bar, in ET, where the threshold appears. If both thresholds hit in one minute/day, -1% is assumed first.",
        ]
    )
    ws.append([])

    headers = [
        "Ticker",
        "Rank Date",
        "Rank",
        "Entry Date",
        "Entry Time",
        "Entry Price",
        "Shares",
        "Investment",
        "Exit Date",
        "Exit Time (ET)",
        "Exit Price",
        "Exit Reason",
        "Status",
        "Data Source",
        "Market Regime",
        "Entry Allowed",
        "Market Reason",
        "Result $",
        "Result %",
    ]
    ws.append(headers)
    table_header_row = ws.max_row
    for row in simulation_rows:
        ws.append(
            [
                row["symbol"],
                row["rank_date"],
                row["rank"],
                row["entry_date"],
                row["entry_time"],
                row["entry_price"],
                row["shares"],
                row["investment"],
                row["exit_date"],
                row["exit_time"],
                row["exit_price"],
                row["exit_reason"],
                row["status"],
                row["data_source"],
                row["market_regime"],
                row["market_entry_allowed"],
                row["market_reason"],
                row["result_currency"],
                row["result_pct"],
            ]
        )

    header_fill = PatternFill(fill_type="solid", fgColor=Color(indexed=8))
    header_font = Font(name="Calibri", size=11, bold=True, color=Color(indexed=9))
    descriptor_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    positive_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    negative_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    blocked_fill = PatternFill(fill_type="solid", fgColor="E7E6E6")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    ws["A1"].value = "Overall Portfolio Standings"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    summary_start_row = 2
    summary_end_row = summary_start_row + len(summary_rows) - 1
    note_row = summary_end_row + 2

    for row_idx in range(summary_start_row, summary_end_row + 1):
        ws.cell(row=row_idx, column=1).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=row_idx, column=1).fill = descriptor_fill
        ws.cell(row=row_idx, column=2).alignment = left_align

    ws.cell(row=note_row, column=1).font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=note_row, column=1).fill = descriptor_fill
    ws.cell(row=note_row, column=2).alignment = left_align
    ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=len(headers))

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=table_header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    currency_format = '"$"#,##0.00;[Red]-"$"#,##0.00'
    pct_format = '0.00%;[Red]-0.00%'
    date_cols = {2, 4, 9}
    currency_cols = {6, 8, 11, 18}
    for row_idx in range(table_header_row + 1, ws.max_row + 1):
        result_value = _coerce_float(ws.cell(row=row_idx, column=18).value)
        status_value = str(ws.cell(row=row_idx, column=13).value or "").strip()
        if status_value == "Blocked":
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = blocked_fill
        elif result_value is not None:
            row_fill = positive_fill if result_value >= 0 else negative_fill
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = row_fill
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = center_align
        for col_idx in date_cols:
            ws.cell(row=row_idx, column=col_idx).number_format = "mmm d, yyyy"
        for col_idx in currency_cols:
            ws.cell(row=row_idx, column=col_idx).number_format = currency_format
        ws.cell(row=row_idx, column=7).number_format = "0.0000"
        ws.cell(row=row_idx, column=19).number_format = pct_format

    summary_currency_rows = {
        summary_label_rows[label]
        for label in {
            "Initial Portfolio Capital",
            "Position Size Per Entry",
            "Open Position Cost",
            "Realized P/L",
            "Portfolio Value (Realized)",
            "Max Capital Deployed",
        }
        if label in summary_label_rows
    }
    for row_idx in summary_currency_rows:
        ws.cell(row=row_idx, column=2).number_format = currency_format
    for label in ("Portfolio Return (Realized)", "SPY 5D Return", "QQQ 5D Return"):
        row_idx = summary_label_rows.get(label)
        if row_idx is not None:
            ws.cell(row=row_idx, column=2).number_format = pct_format
    regime_date_row = summary_label_rows.get("Market Regime Date")
    if regime_date_row is not None:
        ws.cell(row=regime_date_row, column=2).number_format = "mmm d, yyyy"

    ws.freeze_panes = f"A{table_header_row + 1}"
    auto_size_columns(ws, min_width=10, max_width=35)
    ws.column_dimensions["B"].width = 18
    write_summary_sheet(wb, simulation_rows, include_market_status=True)


def remove_inactive_report_sheets(wb: Workbook) -> None:
    for sheet_name in (TOP10_OHLC_SHEET_NAME, INVESTMENT_DASHBOARD_SHEET_NAME):
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])


def write_summary_only_sheet(
    wb: Workbook,
    symbol_paths: dict[str, Path],
    root: Path,
    top_n: int = 10,
    position_size: float = 10_000.0,
    polygon_api_key: str = "",
    intraday_exit_source: str = "auto",
    market_regime_mode: str = MARKET_REGIME_DEFAULT_MODE,
) -> None:
    cohorts = collect_top_ranked_cohorts(wb, top_n=top_n)
    market_cache: dict[str, tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]] = {}
    market_regimes = {
        rank_date: evaluate_market_regime(rank_date, symbol_paths, root, market_cache, mode=market_regime_mode)
        for rank_date in {cohort["rank_date"] for cohort in cohorts}
    }
    simulation_rows = build_investment_simulation_rows(
        cohorts,
        symbol_paths,
        root,
        position_size=position_size,
        polygon_api_key=polygon_api_key,
        intraday_exit_source=intraday_exit_source,
        market_regimes=market_regimes,
    )
    write_summary_sheet(wb, simulation_rows, include_market_status=True)

    am_simulation_rows = build_investment_simulation_rows(
        cohorts,
        symbol_paths,
        root,
        position_size=position_size,
        polygon_api_key=polygon_api_key,
        intraday_exit_source=intraday_exit_source,
        market_regimes=market_regimes,
        entry_session="am",
    )
    write_summary_sheet(
        wb,
        am_simulation_rows,
        sheet_name=AM_SIMULATION_SHEET_NAME,
        include_entry_time=True,
        include_market_status=True,
    )

    pm_simulation_rows = build_investment_simulation_rows(
        cohorts,
        symbol_paths,
        root,
        position_size=position_size,
        polygon_api_key=polygon_api_key,
        intraday_exit_source=intraday_exit_source,
        market_regimes=market_regimes,
        entry_session="pm",
    )
    write_summary_sheet(
        wb,
        pm_simulation_rows,
        sheet_name=PM_SIMULATION_SHEET_NAME,
        include_entry_time=True,
        include_market_status=True,
    )


def write_summary_sheet(
    wb: Workbook,
    simulation_rows: list[dict[str, Any]],
    sheet_name: str = SUMMARY_SHEET_NAME,
    include_entry_time: bool = False,
    include_market_status: bool = False,
) -> None:
    if LEGACY_SUMMARY_SHEET_NAME in wb.sheetnames and SUMMARY_SHEET_NAME not in wb.sheetnames:
        wb[LEGACY_SUMMARY_SHEET_NAME].title = SUMMARY_SHEET_NAME
    elif LEGACY_SUMMARY_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[LEGACY_SUMMARY_SHEET_NAME])

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for merged_range in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged_range))
        ws.delete_rows(1, ws.max_row)
    else:
        insert_at = (
            wb.sheetnames.index(INVESTMENT_DASHBOARD_SHEET_NAME)
            if INVESTMENT_DASHBOARD_SHEET_NAME in wb.sheetnames
            else len(wb.sheetnames)
        )
        if sheet_name == AM_SIMULATION_SHEET_NAME and SUMMARY_SHEET_NAME in wb.sheetnames:
            insert_at = wb.sheetnames.index(SUMMARY_SHEET_NAME) + 1
        elif sheet_name == PM_SIMULATION_SHEET_NAME:
            # Keep the three simulations adjacent and in session order.
            anchor = next(
                (
                    name
                    for name in (AM_SIMULATION_SHEET_NAME, SUMMARY_SHEET_NAME)
                    if name in wb.sheetnames
                ),
                None,
            )
            if anchor is not None:
                insert_at = wb.sheetnames.index(anchor) + 1
        ws = wb.create_sheet(title=sheet_name, index=insert_at)

    headers = [
        "Symbol",
        "Rank \nDate",
        "Entry \nDate",
        "Entry\nPrice",
    ]
    if include_entry_time:
        headers.append("Entry\nTime")
    headers += [
        "Exit\nDate",
        "Exit\nTime",
        "Exit\nPrice",
        "Result\n$",
        "Result\n%",
        "# of Days open",
    ]
    if include_market_status:
        headers.append("SPY - Market Condition")
    ws.append(headers)

    summary_rows = []
    for row in simulation_rows:
        entry_date = row.get("entry_date")
        exit_date = row.get("exit_date")
        if not isinstance(entry_date, date):
            entry_date = None
        if not isinstance(exit_date, date):
            exit_date = None
        status = str(row.get("status", "")).strip()
        is_closed = exit_date is not None
        output_row = [
            row.get("symbol"),
            row.get("rank_date"),
            entry_date,
            row.get("entry_price"),
        ]
        if include_entry_time:
            output_row.append(row.get("entry_time"))
        output_row += [
            exit_date,
            row.get("exit_time") if exit_date else None,
            row.get("exit_price") if exit_date else None,
            row.get("result_currency") if exit_date else None,
            row.get("result_pct") if exit_date else None,
            _trading_days_open(entry_date, exit_date) if is_closed else None,
        ]
        if include_market_status:
            if status in {"Blocked", "Ignored"}:
                condition = row.get("market_reason")
            else:
                condition = row.get("entry_fallback_reason") or "Good"
            output_row.append(condition)
        summary_rows.append(output_row)

    for output_row in summary_rows:
        ws.append(output_row)

    data_start_row = 2
    data_end_row = data_start_row + len(summary_rows) - 1
    total_label_row = data_end_row + 3 if summary_rows else 4
    total_formula_row = total_label_row + 1
    shift = 1 if include_entry_time else 0
    exit_date_col = 5 + shift
    exit_price_col = 7 + shift
    result_currency_col = 8 + shift
    result_pct_col = 9 + shift
    days_open_col = 10 + shift
    last_col = days_open_col
    if include_market_status:
        last_col += 1
    ws.cell(row=total_label_row, column=result_currency_col, value="TOTAL")
    ws.cell(row=total_label_row, column=result_pct_col, value="TOTAL")
    if summary_rows:
        result_currency_letter = get_column_letter(result_currency_col)
        result_pct_letter = get_column_letter(result_pct_col)
        if include_market_status:
            condition_letter = get_column_letter(last_col)
            currency_formula = (
                f'=SUMIF({condition_letter}{data_start_row}:{condition_letter}{data_end_row},"Good*",'
                f'{result_currency_letter}{data_start_row}:{result_currency_letter}{data_end_row})'
            )
            pct_formula = (
                f'=SUMIF({condition_letter}{data_start_row}:{condition_letter}{data_end_row},"Good*",'
                f'{result_pct_letter}{data_start_row}:{result_pct_letter}{data_end_row})'
            )
        else:
            currency_formula = f"=SUM({result_currency_letter}{data_start_row}:{result_currency_letter}{data_end_row})"
            pct_formula = f"=SUM({result_pct_letter}{data_start_row}:{result_pct_letter}{data_end_row})"
        ws.cell(
            row=total_formula_row,
            column=result_currency_col,
            value=currency_formula,
        )
        ws.cell(
            row=total_formula_row,
            column=result_pct_col,
            value=pct_formula,
        )
    else:
        ws.cell(row=total_formula_row, column=result_currency_col, value=0)
        ws.cell(row=total_formula_row, column=result_pct_col, value=0)

    black = Color(indexed=8)
    white = Color(indexed=9)
    light_gray = Color(rgb="D9D9D9")

    black_fill = PatternFill(fill_type="solid", fgColor=black)
    light_gray_fill = PatternFill(fill_type="solid", fgColor=light_gray)
    white_fill = PatternFill(fill_type="solid", fgColor=white)
    header_font = Font(name="Calibri", size=11, bold=True, color=white)
    result_header_font = Font(name="Calibri", size=11, bold=True, color=black)
    symbol_font = Font(name="Calibri", size=13, bold=True, color=black)
    regular_font = Font(name="Calibri", size=11, color=black)
    result_font = Font(name="Calibri", size=13, color=black)
    total_font = Font(name="Calibri", size=13, bold=True, color=black)
    center = Alignment(horizontal="center", vertical="center")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_black = Side(style="thin", color=black)
    thin_light_gray = Side(style="thin", color=light_gray)
    thick_black = Side(style="thick", color=black)

    column_widths = {
        "A": 9.0,
        "B": 12.3125,
        "C": 13.9531,
        "D": 9.65625,
        "E": 13.1953,
        "F": 7.82812,
        "G": 11.5391,
        "H": 13.0,
        "I": 13.0,
        "J": 13.0,
        "K": 13.0,
        "L": 13.0,
        "M": 13.0,
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 37.4
    for row_idx in range(2, total_formula_row + 1):
        ws.row_dimensions[row_idx].height = 19.05

    date_format = "mmm d, yyyy"
    currency_format = '"$"#,##0.00'
    result_currency_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
    result_pct_format = '0%_);[Red]\\(0%\\)'

    def summary_border(col_idx: int, row_idx: int, include_top: bool, include_bottom: bool) -> Border:
        left = thin_black
        right = thin_black
        if col_idx == 1:
            left = thin_light_gray
        if col_idx == last_col:
            right = thin_light_gray
        entry_group_end = 4 + shift
        thick_left_cols = {3, exit_date_col, result_currency_col, days_open_col}
        thick_right_cols = {2, entry_group_end, exit_price_col, result_pct_col}
        if col_idx in thick_left_cols:
            left = thick_black
        if col_idx in thick_right_cols:
            right = thick_black
        top = thin_black if include_top else None
        bottom = thin_black if include_bottom else None
        if row_idx == total_formula_row and col_idx in {result_currency_col, result_pct_col}:
            bottom = thick_black
        return Border(left=left, right=right, top=top, bottom=bottom)

    for col_idx in range(1, last_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = light_gray_fill if col_idx in {result_currency_col, result_pct_col} else black_fill
        cell.font = result_header_font if col_idx in {result_currency_col, result_pct_col} else header_font
        cell.alignment = header_align
        cell.number_format = "@"
        cell.border = summary_border(col_idx, 1, include_top=False, include_bottom=False)

    for row_idx in range(2, total_formula_row + 1):
        for col_idx in range(1, last_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = light_gray_fill if col_idx in {result_currency_col, result_pct_col} else white_fill
            cell.alignment = center
            cell.font = regular_font
            include_top = row_idx > 2 and row_idx <= max(data_end_row, 1)
            include_bottom = row_idx <= max(data_end_row, 1)
            if col_idx == 1:
                cell.font = symbol_font
                cell.number_format = "@"
            elif col_idx in {2, 3, exit_date_col}:
                cell.number_format = date_format
            elif col_idx in {4, exit_price_col}:
                cell.number_format = currency_format
            elif col_idx == result_currency_col:
                cell.font = total_font if row_idx == total_label_row else result_font
                cell.number_format = "@" if row_idx == total_label_row else result_currency_format
            elif col_idx == result_pct_col:
                cell.font = total_font if row_idx == total_label_row else result_font
                cell.number_format = "@" if row_idx == total_label_row else result_pct_format
            else:
                cell.number_format = "General"
            if row_idx > data_end_row and row_idx not in {total_label_row, total_formula_row}:
                include_top = row_idx == data_end_row + 1
                include_bottom = False
            cell.border = summary_border(col_idx, row_idx, include_top=include_top, include_bottom=include_bottom)

    if include_market_status:
        market_condition_col = last_col
        ws.column_dimensions[get_column_letter(market_condition_col)].width = 42.0
        red_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
        red_font = Font(name="Calibri", size=11, bold=True, color="9C0006")
        ignored_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        ignored_font = Font(name="Calibri", size=11, bold=True, color="7F6000")
        for row_idx, simulation_row in enumerate(simulation_rows, start=2):
            row_status = str(simulation_row.get("status", "")).strip()
            used_regular_fallback = bool(simulation_row.get("entry_fallback_reason"))
            if row_status not in {"Blocked", "Ignored"} and not used_regular_fallback:
                continue
            row_fill = red_fill if row_status == "Blocked" else ignored_fill
            for col_idx in range(1, last_col + 1):
                ws.cell(row=row_idx, column=col_idx).fill = row_fill
            ws.cell(row=row_idx, column=market_condition_col).font = (
                red_font if row_status == "Blocked" else ignored_font
            )

    ws.freeze_panes = None


def prune_old_run_sheets(
    wb: Workbook,
    keep_runs: int = 7,
    protected_names: set[str] | None = None,
) -> None:
    """
    Keep only the most recent run-result sheets in workbook order.
    """
    if keep_runs < 0:
        keep_runs = 0
    keep_protected = protected_names or PROTECTED_SHEET_NAMES
    run_sheet_names = [name for name in wb.sheetnames if name not in keep_protected]
    permanently_retained = {
        name
        for name in run_sheet_names
        if (_parse_run_sheet_date(name) or date.min) >= SIMULATION_START_DATE
    }
    removable_names = [name for name in run_sheet_names if name not in permanently_retained]
    remove_count = len(removable_names) - keep_runs
    if remove_count <= 0:
        return
    for name in removable_names[:remove_count]:
        del wb[name]


def auto_size_columns(ws, min_width: int = 8, max_width: int = 40) -> None:
    """
    Auto-size worksheet column widths based on cell contents.
    """
    for col_idx, col_cells in enumerate(ws.iter_cols(min_col=1, max_col=ws.max_column), start=1):
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            value_str = str(cell.value)
            if "\n" in value_str:
                value_len = max(len(line) for line in value_str.splitlines())
            else:
                value_len = len(value_str)
            if value_len > max_len:
                max_len = value_len
        if max_len == 0:
            continue
        width = min(max_width, max(min_width, max_len + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


_WORKER_PARAMS: dict[str, Any] = {}
_BENCH_MAP: dict[int, float] = {}
_NEED_ROWS = 0


def _init_worker(
    bench_map: dict[int, float],
    params: dict[str, Any],
    need_rows: int,
    as_of_date_int: int | None,
) -> None:
    global _WORKER_PARAMS, _BENCH_MAP, _NEED_ROWS, AS_OF_DATE_INT
    _WORKER_PARAMS = params
    _BENCH_MAP = bench_map
    _NEED_ROWS = need_rows
    AS_OF_DATE_INT = as_of_date_int


def screen_symbol(
    sym: str,
    path: Path,
    params: dict[str, Any],
    bench_map: dict[int, float],
    need_rows: int,
    apply_filters: bool = True,
) -> dict[str, Any] | None:
    d, c, v, h, l = load_series_from_file(path, need_rows=need_rows)
    if len(d) == 0:
        return None
    min_rows = 60
    if params["avg_vol_mode"] == "days":
        min_rows = max(params["avg_vol_days"] + 1, min_rows)
    min_rows = max(min_rows, params["rsi_period"] + 2, params["atr_period"] + 2)
    if apply_filters and len(d) < min_rows:
        return None

    change_lookback = 5  # trading periods
    prev_idx = -(change_lookback + 1)

    # Average daily $ volume filter (close * volume) over avg volume window
    if params["avg_vol_mode"] == "months":
        last_date = date_from_int(int(d[-1]))
        cutoff_date = shift_months(last_date, -int(params["avg_vol_months"]))
        cutoff_int = date_to_int(cutoff_date)
        if apply_filters and d[0] > cutoff_int:
            return None
        mask = d >= cutoff_int
        if apply_filters and not np.any(mask):
            return None
        if np.any(mask):
            close_window = c[mask]
            vol_window = v[mask]
        else:
            close_window = c
            vol_window = v
    else:
        if params["avg_vol_days"] > 0:
            close_window = c[-params["avg_vol_days"]:]
            vol_window = v[-params["avg_vol_days"]:]
        else:
            close_window = c
            vol_window = v

    avg_dollar_volume = float(np.mean(close_window * vol_window)) if len(close_window) else np.nan
    if apply_filters and (not np.isfinite(avg_dollar_volume) or avg_dollar_volume <= params["avg_dollar_vol_min"]):
        return None

    avg_dollar_volume_prev = np.nan
    if len(d) >= change_lookback + 1:
        if params["avg_vol_mode"] == "months":
            last_date_prev = date_from_int(int(d[prev_idx]))
            cutoff_date_prev = shift_months(last_date_prev, -int(params["avg_vol_months"]))
            cutoff_int_prev = date_to_int(cutoff_date_prev)
            last_date_prev_int = date_to_int(last_date_prev)
            mask_prev = (d >= cutoff_int_prev) & (d <= last_date_prev_int)
            if np.any(mask_prev):
                avg_dollar_volume_prev = float(np.mean(c[mask_prev] * v[mask_prev]))
        else:
            lookback_days = int(params["avg_vol_days"])
            if len(c) >= lookback_days + change_lookback:
                close_window_prev = c[-(lookback_days + change_lookback):-change_lookback]
                vol_window_prev = v[-(lookback_days + change_lookback):-change_lookback]
                if len(close_window_prev) and len(vol_window_prev):
                    avg_dollar_volume_prev = float(np.mean(close_window_prev * vol_window_prev))

    last_close = float(c[-1])
    last_date = date_from_int(int(d[-1]))
    prev_close = float(c[prev_idx]) if len(c) >= change_lookback + 1 else np.nan

    cutoff_52_date = last_date - timedelta(days=364)
    cutoff_52_int = date_to_int(cutoff_52_date)
    high_52, high_52_days = max_value_and_days_ago(d, h, last_date, cutoff_52_int)
    # Full-file scan: d/h only cover the tail window, which is roughly one year.
    history_high = scan_all_time_high(path)
    if history_high is None:
        high_all, high_all_days = max_value_and_days_ago(d, h, last_date, None)
    else:
        high_all = history_high.high
        high_all_days = (last_date - date_from_int(history_high.date_int)).days
    last_5pct_date, last_5pct_days = last_close_5pct_higher_info(d, c, last_close, last_date)

    # RSI filter
    rsi_vals = rsi_wilder(c, period=params["rsi_period"])
    last_rsi = float(rsi_vals[-1])
    prev_rsi = float(rsi_vals[prev_idx]) if len(rsi_vals) >= change_lookback + 1 else np.nan
    if apply_filters and not (params["rsi_low"] <= last_rsi <= params["rsi_high"]):
        return None

    # ATR filter (percent of last close)
    atr_vals = atr_wilder(h, l, c, period=params["atr_period"])
    last_atr = float(atr_vals[-1])
    atr_pct = np.nan
    if np.isfinite(last_atr) and last_close > 0:
        atr_pct = (last_atr / last_close) * 100.0
    if apply_filters and (not np.isfinite(atr_pct) or atr_pct <= params["atr_min_pct"]):
        return None

    # MACD filter
    macd_line, sig_line = macd(c, params["macd_fast"], params["macd_slow"], params["macd_signal"])
    last_macd = float(macd_line[-1])
    last_sig = float(sig_line[-1])
    prev_macd = float(macd_line[prev_idx]) if len(macd_line) >= change_lookback + 1 else np.nan
    prev_sig = float(sig_line[prev_idx]) if len(sig_line) >= change_lookback + 1 else np.nan
    if apply_filters and not (last_macd > last_sig):
        return None

    macd_signal_ratio = np.nan
    if np.isfinite(last_macd) and np.isfinite(last_sig) and last_sig != 0:
        macd_signal_ratio = last_macd / last_sig
    macd_signal_ratio_prev = np.nan
    if np.isfinite(prev_macd) and np.isfinite(prev_sig) and prev_sig != 0:
        macd_signal_ratio_prev = prev_macd / prev_sig

    stock_close_aligned = np.array([], dtype=float)
    bench_close_aligned = np.array([], dtype=float)
    beta_prev = np.nan

    if params["beta_freq"] == "monthly":
        sm, smc = monthly_closes(d, c)
        if len(sm) < params["beta_months"] + 1:
            if apply_filters:
                return None
        else:
            stock_aligned: list[float] = []
            bench_aligned: list[float] = []
            for mk, ci in zip(sm, smc):
                bench_ci = bench_map.get(int(mk))
                if bench_ci is None:
                    continue
                stock_aligned.append(float(ci))
                bench_aligned.append(float(bench_ci))

            if len(stock_aligned) < params["beta_months"] + 1:
                if apply_filters:
                    return None
            else:
                stock_close_aligned = np.array(stock_aligned[-(params["beta_months"] + 1):], dtype=float)
                bench_close_aligned = np.array(bench_aligned[-(params["beta_months"] + 1):], dtype=float)
                if len(stock_aligned) >= params["beta_months"] + 1 + change_lookback:
                    prev_stock = np.array(stock_aligned[:-change_lookback], dtype=float)
                    prev_bench = np.array(bench_aligned[:-change_lookback], dtype=float)
                    if len(prev_stock) >= params["beta_months"] + 1:
                        beta_prev = beta_from_aligned_closes(
                            prev_stock[-(params["beta_months"] + 1):],
                            prev_bench[-(params["beta_months"] + 1):],
                        )
    else:
        # Align stock closes to benchmark dates for beta (no dict/sort needed)
        stock_aligned = []
        bench_aligned = []
        for di, ci in zip(d, c):
            bench_ci = bench_map.get(int(di))
            if bench_ci is None:
                continue
            stock_aligned.append(float(ci))
            bench_aligned.append(float(bench_ci))

        if len(stock_aligned) < params["beta_lookback"] + 1:
            if apply_filters:
                return None
        else:
            stock_close_aligned = np.array(stock_aligned[-(params["beta_lookback"] + 1):], dtype=float)
            bench_close_aligned = np.array(bench_aligned[-(params["beta_lookback"] + 1):], dtype=float)
            if len(stock_aligned) >= params["beta_lookback"] + 1 + change_lookback:
                prev_stock = np.array(stock_aligned[:-change_lookback], dtype=float)
                prev_bench = np.array(bench_aligned[:-change_lookback], dtype=float)
                if len(prev_stock) >= params["beta_lookback"] + 1:
                    beta_prev = beta_from_aligned_closes(
                        prev_stock[-(params["beta_lookback"] + 1):],
                        prev_bench[-(params["beta_lookback"] + 1):],
                    )

    b = beta_from_aligned_closes(stock_close_aligned, bench_close_aligned)
    if apply_filters and (not np.isfinite(b) or b <= params["beta_min"]):
        return None

    close_pct_5 = pct_change(last_close, prev_close)
    beta_pct_5 = pct_change(b, beta_prev)
    rsi_pct_5 = pct_change(last_rsi, prev_rsi)
    macd_pct_5 = pct_change(last_macd, prev_macd)
    signal_pct_5 = pct_change(last_sig, prev_sig)
    macd_signal_ratio_pct_5 = pct_change(macd_signal_ratio, macd_signal_ratio_prev)
    avg_dollar_volume_pct_5 = pct_change(avg_dollar_volume, avg_dollar_volume_prev)

    return {
        "symbol": display_symbol(sym),  # NO ".US" in output
        "last_close": last_close,
        "last_close_pct_5": close_pct_5,
        "high_52w_close": high_52,
        "high_52w_days_ago": high_52_days,
        "high_all_close": high_all,
        "high_all_days_ago": high_all_days,
        "last_5pct_higher_date": last_5pct_date,
        "last_5pct_higher_days_ago": last_5pct_days,
        "beta": b,
        "beta_pct_5": beta_pct_5,
        "rsi": last_rsi,
        "rsi_pct_5": rsi_pct_5,
        "atr": last_atr,
        "atr_pct": atr_pct,
        "macd": last_macd,
        "macd_pct_5": macd_pct_5,
        "signal": last_sig,
        "signal_pct_5": signal_pct_5,
        "macd_signal_ratio": macd_signal_ratio,
        "macd_signal_ratio_pct_5": macd_signal_ratio_pct_5,
        "avg_dollar_volume": avg_dollar_volume,
        "avg_dollar_volume_pct_5": avg_dollar_volume_pct_5,
    }


def _screen_symbol_worker(task: tuple[str, Path]) -> dict[str, Any] | None:
    sym, path = task
    return screen_symbol(sym, path, _WORKER_PARAMS, _BENCH_MAP, _NEED_ROWS)


# -----------------------------
# Main screener
# -----------------------------
def main() -> None:
    ap = argparse.ArgumentParser()
    tickers_group = ap.add_mutually_exclusive_group(required=False)
    tickers_group.add_argument("--tickers", help="CSV with tickers (column: symbol or ticker)")
    tickers_group.add_argument(
        "--tickers_dir",
        help='Directory containing *.us.txt files to screen (e.g. "${workspaceFolder}/data/daily/us/nyse stocks")',
    )
    ap.add_argument("--root", required=True, help='Root folder: e.g. "${workspaceFolder}/data/daily/us" or "/Users/v/Downloads/data/daily/us"')
    ap.add_argument(
        "--out",
        default="results.xlsx",
        help='Output Excel (.xlsx) path (supports ${workspaceFolder}, ~, env vars)',
    )
    ap.add_argument(
        "--as_of_date",
        default="",
        help="Run the screener using data available through YYYY-MM-DD (for historical backfills).",
    )
    ap.add_argument(
        "--alphavantage_api_key",
        default=ALPHAVANTAGE_API_KEY,
        help="Alpha Vantage API key (defaults to hardcoded project key).",
    )
    ap.add_argument(
        "--splits_lookback_days",
        type=int,
        default=90,
        help="How far back the Recent Splits sheet looks (default: 90).",
    )
    ap.add_argument(
        "--polygon_api_key",
        default=os.environ.get("POLYGON_API_KEY", ""),
        help="Polygon API key for 1-minute investment dashboard exits. Defaults to POLYGON_API_KEY.",
    )
    ap.add_argument(
        "--intraday_exit_source",
        choices=["auto", "polygon", "daily"],
        default="auto",
        help="Dashboard exit data source: auto/polygon uses Polygon 1-minute bars when a key is available; daily uses daily OHLC.",
    )
    ap.add_argument(
        "--market_regime_mode",
        choices=MARKET_REGIME_MODES,
        default=MARKET_REGIME_DEFAULT_MODE,
        help=(
            # Percent signs are doubled: argparse runs help through %-formatting,
            # and Python 3.14 rejects a single %% as a badly formed help string.
            "Market-regime entry gate for the Investment Dashboard. "
            "standard uses SPY > 50DMA and SPY 5D > -2%%; "
            "aggressive also requires SPY > 200DMA, SPY 20DMA > 50DMA, and SPY 5D > 0%%."
        ),
    )

    ap.add_argument("--benchmark", default="SPY.US")

    ap.add_argument(
        "--beta_lookback",
        type=int,
        default=252,
        help="Beta lookback in trading days (default: 252). Used when --beta_freq daily.",
    )
    ap.add_argument(
        "--beta_freq",
        choices=["daily", "monthly"],
        default="daily",
        help="Beta frequency: daily or monthly (default: daily)",
    )
    ap.add_argument(
        "--beta_months",
        type=int,
        default=60,
        help="Beta lookback in months (default: 60). Used when --beta_freq monthly.",
    )
    ap.add_argument("--beta_min", type=float, default=1.2, help="Minimum beta (default: 1.2)")

    ap.add_argument("--rsi_low", type=float, default=50.0)
    ap.add_argument("--rsi_high", type=float, default=70.0)
    ap.add_argument("--rsi_period", type=int, default=14)

    ap.add_argument("--atr_period", type=int, default=14, help="ATR period in days (default: 14)")
    ap.add_argument(
        "--atr_min_pct",
        type=float,
        default=2.0,
        help="Minimum ATR as percent of last close (default: 2.0)",
    )

    ap.add_argument("--macd_fast", type=int, default=12)
    ap.add_argument("--macd_slow", type=int, default=26)
    ap.add_argument("--macd_signal", type=int, default=12)

    ap.add_argument(
        "--avg_vol_days",
        type=int,
        default=None,
        help="Window for avg volume (trading days). Overrides --avg_vol_months when set.",
    )
    ap.add_argument(
        "--avg_vol_months",
        type=int,
        default=6,
        help="Window for avg volume (calendar months) (default: 6)",
    )

    ap.add_argument(
        "--workers",
        type=int,
        default=0,
        help="Parallel workers for symbol screening (0=auto, 1=disable parallelism)",
    )
    ap.add_argument(
        "--run_mode",
        choices=["all", "single"],
        default="",
        help="Run mode for unattended use. Omit to show the interactive menu.",
    )
    ap.add_argument(
        "--single_symbol",
        default="",
        help="Ticker to screen when --run_mode single is used, e.g. AAPL or AAPL.US.",
    )

    # Average daily dollar volume filter (close * volume)
    ap.add_argument(
        "--avg_dollar_vol_min",
        type=float,
        default=5_000_000.0,
        help="Minimum average daily $ volume over avg volume window (close * volume). Default: 5,000,000",
    )
    ap.add_argument(
        "--top_n",
        type=int,
        default=10,
        help="Number of top-ranked rows to flag (default: 10)",
    )
    ap.add_argument(
        "--daily_limit",
        type=int,
        default=25,
        help="Maximum rows to write on each all-tickers daily results sheet (default: 25)",
    )
    ap.add_argument(
        "--score_enable",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Enable Top-N scoring and ranking columns (default: True)",
    )

    args = ap.parse_args()

    global AS_OF_DATE_INT
    if args.as_of_date:
        try:
            AS_OF_DATE_INT = int(datetime.strptime(args.as_of_date, "%Y-%m-%d").strftime("%Y%m%d"))
        except ValueError as exc:
            raise SystemExit("--as_of_date must use YYYY-MM-DD.") from exc

    if args.avg_vol_days is not None and args.avg_vol_days <= 0:
        raise SystemExit("--avg_vol_days must be > 0")
    if args.avg_vol_months <= 0:
        raise SystemExit("--avg_vol_months must be > 0")
    if args.atr_period <= 0:
        raise SystemExit("--atr_period must be > 0")
    if args.atr_min_pct < 0:
        raise SystemExit("--atr_min_pct must be >= 0")
    if args.daily_limit <= 0:
        raise SystemExit("--daily_limit must be > 0")
    if args.intraday_exit_source == "polygon" and not args.polygon_api_key.strip():
        raise SystemExit("--intraday_exit_source polygon requires --polygon_api_key or POLYGON_API_KEY.")

    if args.run_mode:
        run_mode = args.run_mode
        single_symbol = normalize_symbol(args.single_symbol) if args.single_symbol else None
        if run_mode == "single" and not single_symbol:
            raise SystemExit("--single_symbol is required when --run_mode single is used.")
    else:
        run_mode, single_symbol = prompt_run_mode()

    root = resolve_path(args.root)
    out_path = resolve_path(args.out)

    tickers: list[str] = []
    symbol_paths: dict[str, Path] = {}
    single_path: Path | None = None

    if run_mode == "all":
        if not args.tickers and not args.tickers_dir:
            raise SystemExit("Provide --tickers or --tickers_dir for All tickers mode.")
        if args.tickers:
            tickers_path = resolve_path(args.tickers)
            tickers = load_tickers_csv(tickers_path)
            symbol_paths = build_file_map(root)
        else:
            tickers_root = resolve_path(args.tickers_dir)
            symbol_paths = build_file_map(tickers_root)
            tickers = sorted(symbol_paths.keys())
    else:
        if not single_symbol:
            raise SystemExit("Single ticker selection required.")

    bench_sym = normalize_symbol(args.benchmark)

    bench_path = None
    if run_mode == "all":
        bench_path = symbol_paths.get(bench_sym)
        if bench_path is None:
            bench_path = find_symbol_file(root, bench_sym)
    else:
        bench_path = find_symbol_file(root, bench_sym)
        if bench_path is None and args.tickers_dir:
            bench_path = find_symbol_file(resolve_path(args.tickers_dir), bench_sym)

    if bench_path is None:
        raise SystemExit(
            f"Benchmark {bench_sym} not found under {root}.\n"
            f"Tip: point --root at a parent folder that includes both stocks + ETFs."
        )

    beta_rows = args.beta_lookback + 10
    if args.beta_freq == "monthly":
        beta_rows = args.beta_months * 23 + 10

    if args.avg_vol_days is not None and args.avg_vol_days > 0:
        avg_vol_mode = "days"
        avg_vol_days = args.avg_vol_days
        avg_vol_months = 0
        avg_vol_need_rows = args.avg_vol_days + 30
    else:
        avg_vol_mode = "months"
        avg_vol_days = 0
        avg_vol_months = args.avg_vol_months
        avg_vol_need_rows = args.avg_vol_months * 23 + 30

    # Read enough rows from each file to compute beta + indicators + averages.
    need_rows = max(
        beta_rows,
        args.macd_slow + args.macd_signal + 30,
        avg_vol_need_rows,
        args.atr_period + 30,
        260,
    )

    # Load benchmark series (tail), build date->close map or month->close map
    bd, bc, _bv, _bh, _bl = load_series_from_file(bench_path, need_rows=need_rows)
    if args.beta_freq == "monthly":
        bm, bmc = monthly_closes(bd, bc)
        if len(bm) < args.beta_months + 1:
            raise SystemExit(
                f"Not enough benchmark history for {bench_sym} "
                f"(have {len(bm)} months, need {args.beta_months + 1})."
            )
        bm = bm[-(args.beta_months + 1):]
        bmc = bmc[-(args.beta_months + 1):]
        bench_map = {int(mi): float(ci) for mi, ci in zip(bm, bmc)}
    else:
        if len(bd) < args.beta_lookback + 1:
            raise SystemExit(f"Not enough benchmark history for {bench_sym} (have {len(bd)} rows).")

        bd = bd[-(args.beta_lookback + 10):]
        bc = bc[-(args.beta_lookback + 10):]
        bench_map = {int(di): float(ci) for di, ci in zip(bd, bc)}

    tasks: list[tuple[str, Path]] = []
    if run_mode == "all":
        for sym in tickers:
            p = symbol_paths.get(sym)
            if not p:
                continue
            tasks.append((sym, p))
    else:
        single_path = find_symbol_file(root, single_symbol)
        if single_path is None and args.tickers_dir:
            single_path = find_symbol_file(resolve_path(args.tickers_dir), single_symbol)
        if single_path is None:
            raise SystemExit(f"Ticker {single_symbol} not found under {root}.")
        tasks.append((single_symbol, single_path))

    params = {
        "beta_freq": args.beta_freq,
        "beta_lookback": args.beta_lookback,
        "beta_months": args.beta_months,
        "beta_min": args.beta_min,
        "rsi_low": args.rsi_low,
        "rsi_high": args.rsi_high,
        "rsi_period": args.rsi_period,
        "atr_period": args.atr_period,
        "atr_min_pct": args.atr_min_pct,
        "macd_fast": args.macd_fast,
        "macd_slow": args.macd_slow,
        "macd_signal": args.macd_signal,
        "avg_vol_mode": avg_vol_mode,
        "avg_vol_days": avg_vol_days,
        "avg_vol_months": avg_vol_months,
        "avg_dollar_vol_min": args.avg_dollar_vol_min,
    }

    cpu_count = os.cpu_count() or 1
    workers = args.workers if args.workers > 0 else max(1, cpu_count - 1)
    workers = min(workers, len(tasks)) if tasks else 1

    results: list[dict[str, Any]] = []
    if run_mode == "all":
        if workers > 1 and len(tasks) > 1:
            chunksize = max(1, len(tasks) // (workers * 4))
            with ProcessPoolExecutor(
                max_workers=workers,
                initializer=_init_worker,
                initargs=(bench_map, params, need_rows, AS_OF_DATE_INT),
            ) as ex:
                for res in ex.map(_screen_symbol_worker, tasks, chunksize=chunksize):
                    if res:
                        results.append(res)
        else:
            for sym, p in tasks:
                res = screen_symbol(sym, p, params, bench_map, need_rows)
                if res:
                    results.append(res)
    else:
        sym, p = tasks[0]
        res = screen_symbol(sym, p, params, bench_map, need_rows, apply_filters=False)
        if res:
            results.append(res)
        else:
            raise SystemExit(f"No data available for {display_symbol(sym)}.")

    if results:
        symbols = [str(row.get("symbol", "")).strip().upper() for row in results]
        session = requests.Session()
        earnings_reference_date = date_from_int(AS_OF_DATE_INT) if AS_OF_DATE_INT is not None else date.today()
        earnings_map = fetch_nasdaq_earnings_dates(symbols, session, today=earnings_reference_date)
        for row in results:
            symbol = str(row.get("symbol", "")).strip().upper()
            info = earnings_map.get(symbol, {})
            row["last_earnings_date"] = info.get("last", "")
            row["next_earnings_date"] = info.get("next", "")
            row["company_name"] = info.get("name", "")

    def parse_earnings_date(value: Any) -> date | None:
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value).strip()
        if not text:
            return None
        try:
            return datetime.strptime(text, "%m/%d/%Y").date()
        except Exception:
            return None

    def to_float(value: Any) -> float | None:
        if value is None:
            return None
        try:
            out = float(value)
        except Exception:
            return None
        return out if np.isfinite(out) else None

    def percentile_rank(values: list[float], value: float) -> float:
        if not values:
            return 0.0
        sorted_vals = sorted(values)
        n = len(sorted_vals)
        if n == 1:
            return 1.0
        left = bisect_left(sorted_vals, value)
        right = bisect_right(sorted_vals, value)
        idx = (left + right - 1) / 2.0
        return idx / (n - 1)

    def score_row(row: dict[str, Any], liquidity_values: list[float]) -> float | None:
        avg_dollar_volume = to_float(row.get("avg_dollar_volume"))
        atr_pct = to_float(row.get("atr_pct"))
        rsi = to_float(row.get("rsi"))
        if avg_dollar_volume is None or atr_pct is None or rsi is None:
            return None
        liquidity_score = percentile_rank(liquidity_values, avg_dollar_volume)
        if 3.0 <= atr_pct <= 5.5:
            atr_score = 1.0
        else:
            distance = (3.0 - atr_pct) if atr_pct < 3.0 else (atr_pct - 5.5)
            atr_score = max(0.0, 1.0 - distance / 3.0)
        rsi_score = max(0.0, min(1.0, 1.0 - abs(rsi - 55.0) / 15.0))
        macd_ratio_pct = to_float(row.get("macd_signal_ratio_pct_5"))
        momentum_score = 1.0 if macd_ratio_pct is not None and macd_ratio_pct > 0 else 0.0
        prev_runs = to_float(row.get("prev_5_runs")) or 0.0
        consistency_score = min(prev_runs / 5.0, 1.0)
        return (
            0.30 * liquidity_score
            + 0.25 * atr_score
            + 0.20 * rsi_score
            + 0.15 * momentum_score
            + 0.10 * consistency_score
        )

    def is_score_eligible(row: dict[str, Any], today: date) -> bool:
        next_earnings = parse_earnings_date(row.get("next_earnings_date"))
        if next_earnings is not None:
            days_out = (next_earnings - today).days
            if 0 <= days_out <= 10:
                return False
        atr_pct = to_float(row.get("atr_pct"))
        if atr_pct is None or atr_pct < 2.5:
            return False
        avg_dollar_volume = to_float(row.get("avg_dollar_volume"))
        if avg_dollar_volume is None or avg_dollar_volume < 10_000_000:
            return False
        rsi = to_float(row.get("rsi"))
        if rsi is None or rsi >= 68:
            return False
        return True

    out_path.parent.mkdir(parents=True, exist_ok=True)
    if out_path.suffix.lower() != ".xlsx":
        out_path = out_path.with_suffix(".xlsx")

    # Row 1 spells out what each number means in plain words; row 2 (descriptors,
    # below) carries the settings behind it. The "How It Works" tab explains both.
    header_row = [
        "Symbol",
        "Company",
        "Times Seen",
        "Closing Price",
        None,
        "52-Week High",
        None,
        "2-Year High",
        None,
        "Last 5% Higher Close",
        None,
        "Beta vs S&P 500",
        None,
        "RSI Momentum",
        None,
        "ATR Daily Swing %",
        "MACD Trend",
        None,
        "MACD Signal Line",
        None,
        "MACD vs Signal",
        None,
        "Avg Daily $ Volume",
        None,
        "Earnings Dates",
        None,
    ]
    data_keys = [
        "symbol",
        "company_name",
        "prev_5_runs",
        "last_close",
        "last_close_pct_5",
        "high_52w_close",
        "high_52w_days_ago",
        "high_all_close",
        "high_all_days_ago",
        "last_5pct_higher_date",
        "last_5pct_higher_days_ago",
        "beta",
        "beta_pct_5",
        "rsi",
        "rsi_pct_5",
        "atr_pct",
        "macd",
        "macd_pct_5",
        "signal",
        "signal_pct_5",
        "macd_signal_ratio",
        "macd_signal_ratio_pct_5",
        "avg_dollar_volume",
        "avg_dollar_volume_pct_5",
        "last_earnings_date",
        "next_earnings_date",
    ]
    if args.score_enable:
        header_row = header_row[:2] + ["Rank"] + header_row[2:]
        data_keys = data_keys[:2] + ["rank"] + data_keys[2:]
    header_row = ["Run Date"] + header_row
    data_date = date_from_int(int(bd[-1])) if len(bd) else date.today()
    headline = data_date.strftime("%d %b %Y").upper()

    if out_path.exists():
        wb = load_workbook(out_path)
        if run_mode == "all":
            prepare_daily_runs_sheet(wb, header_row)
        prev_counts = count_recent_symbol_occurrences(wb, max_runs=5, before_date=data_date) if run_mode == "all" else {}
    else:
        wb = Workbook()
        prev_counts = {}

    daily_output_rows: list[dict[str, Any]] = []
    top10: list[dict[str, Any]] = []
    if results:
        for row in results:
            symbol_key = str(row.get("symbol", "")).strip().upper()
            row["prev_5_runs"] = int(prev_counts.get(symbol_key, 0))
            if args.score_enable:
                row["total_score"] = None
                row["rank"] = None
    if results and args.score_enable:
        score_today = data_date
        eligible = [row for row in results if is_score_eligible(row, score_today)]
        liquidity_values = [to_float(row.get("avg_dollar_volume")) for row in eligible]
        liquidity_values = [val for val in liquidity_values if val is not None]
        for row in eligible:
            row["total_score"] = score_row(row, liquidity_values)
        scored = sorted(
            [row for row in eligible if row.get("total_score") is not None],
            key=lambda r: float(r.get("total_score", 0.0)),
            reverse=True,
        )
        for idx, row in enumerate(scored, start=1):
            row["_daily_order"] = idx
        top_n = max(args.top_n, 0)
        top10 = scored[:top_n] if top_n else []
        for idx, row in enumerate(top10, start=1):
            row["rank"] = idx
        daily_output_rows = scored[: args.daily_limit]

    if not daily_output_rows:
        daily_output_rows = sorted(
            results,
            key=lambda r: str(r.get("symbol", "")).strip().upper(),
        )[: args.daily_limit]

    splits_session = requests.Session()

    # Universe-wide sweep: a split corrupts a symbol's stored history whether or
    # not that symbol ranked today, so this is not limited to reported rows.
    splits_start = data_date - timedelta(days=args.splits_lookback_days)
    tracked_symbols = {
        display_symbol(sym).upper()
        for sym in (symbol_paths.keys() if symbol_paths else [t for t in tickers])
    }
    if run_mode == "single":
        tracked_symbols = {display_symbol(single_symbol).upper()}
    write_recent_splits_sheet(
        wb,
        fetch_polygon_splits_since(
            args.polygon_api_key, splits_session, splits_start, tracked_symbols
        ),
        splits_start,
        data_date,
        have_api_key=bool(args.polygon_api_key.strip()),
    )

    qualified_dates = collect_qualified_result_dates(
        wb,
        current_results=daily_output_rows if run_mode == "all" else None,
        current_date=data_date if run_mode == "all" else None,
    )

    if run_mode == "single":
        if "Single Tickers" in wb.sheetnames:
            ws = wb["Single Tickers"]
        elif len(wb.sheetnames) == 1 and is_empty_sheet(wb.active):
            ws = wb.active
            ws.title = "Single Tickers"
        else:
            ws = wb.create_sheet(title="Single Tickers")
    else:
        sheet_name = DAILY_RUNS_SHEET_NAME
        ws = prepare_daily_runs_sheet(wb, header_row) if DAILY_RUNS_SHEET_NAME not in wb.sheetnames else wb[DAILY_RUNS_SHEET_NAME]

    if args.beta_freq == "monthly":
        beta_desc = f"{args.beta_months} months\n> {args.beta_min}"
    else:
        beta_desc = f"{args.beta_lookback} days\n> {args.beta_min}"

    if avg_vol_mode == "months":
        avg_desc = f"{avg_vol_months} months\n> ${args.avg_dollar_vol_min:,.0f}"
    else:
        avg_desc = f"{avg_vol_days} days\n> ${args.avg_dollar_vol_min:,.0f}"

    pct_change_days = 5
    pct_change_desc = f"vs {pct_change_days}\ndays ago"
    pct_desc = f"vs {pct_change_days} days ago"
    beta_change_desc = (
        f"vs {pct_change_days}\nmonths ago" if args.beta_freq == "monthly" else pct_change_desc
    )
    descriptors = [
        None,
        None,
        "of the last\n5 runs",
        "Latest close",
        pct_change_desc,
        "Best close\nin 1 year",
        "Days ago",
        "Best close\nin the data",
        "Days ago",
        "Date",
        "Days ago",
        beta_desc,
        beta_change_desc,
        f"{args.rsi_period} days\n{args.rsi_low} to {args.rsi_high}",
        pct_desc,
        f"{args.atr_period} days\n> {args.atr_min_pct}% of close",
        f"{args.macd_fast}/{args.macd_slow} EMA\nMACD > Signal",
        pct_desc,
        f"{args.macd_signal} days",
        pct_desc,
        "MACD / Signal",
        pct_desc,
        avg_desc,
        pct_desc,
        "Most recent",
        "Next",
    ]
    if args.score_enable:
        descriptors = descriptors[:2] + [None] + descriptors[2:]
    descriptors = [None] + descriptors
    if run_mode == "all":
        for col_idx, value in enumerate(descriptors, start=1):
            ws.cell(row=2, column=col_idx, value=value)

    def parse_mmddyyyy(value: Any) -> datetime | None:
        if not value:
            return None
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime.combine(value, datetime.min.time())
        try:
            return datetime.strptime(str(value), "%m/%d/%Y")
        except Exception:
            return None

    def fmt_pct_value(x: Any) -> float | None:
        if x is None:
            return None
        try:
            if isinstance(x, (float, np.floating)) and not np.isfinite(x):
                return None
            return float(f"{float(x):.2f}")
        except Exception:
            return None

    def build_output_row(row: dict[str, Any], prev_count: int | None) -> list[Any]:
        high_52_close = row.get("high_52w_close")
        high_all_close = row.get("high_all_close")
        symbol_display = str(row.get("symbol", "")).strip()
        company_name = str(row.get("company_name", "")).strip()
        if company_name:
            company_name = company_name[:10]
        last_close_val = row.get("last_close")
        last_close = float(last_close_val) if last_close_val is not None and np.isfinite(last_close_val) else None
        avg_dollar_volume = row.get("avg_dollar_volume")
        avg_dollar_volume_val = (
            float(avg_dollar_volume) if avg_dollar_volume is not None and np.isfinite(avg_dollar_volume) else None
        )
        rank_val = row.get("rank") if args.score_enable else None
        output_row = [
            symbol_display,
            company_name,
        ]
        if args.score_enable:
            output_row.append(rank_val)
        output_row.extend(
            [
                prev_count,
                last_close,
                fmt_pct_value(row.get("last_close_pct_5")),
                float(high_52_close) if high_52_close is not None else None,
                row.get("high_52w_days_ago"),
                float(high_all_close) if high_all_close is not None else None,
                row.get("high_all_days_ago"),
                parse_mmddyyyy(row.get("last_5pct_higher_date")),
                row.get("last_5pct_higher_days_ago"),
                to_float(row.get("beta")),
                fmt_pct_value(row.get("beta_pct_5")),
                to_float(row.get("rsi")),
                fmt_pct_value(row.get("rsi_pct_5")),
                fmt_pct_value(row.get("atr_pct")),
                to_float(row.get("macd")),
                fmt_pct_value(row.get("macd_pct_5")),
                to_float(row.get("signal")),
                fmt_pct_value(row.get("signal_pct_5")),
                fmt2(row.get("macd_signal_ratio")),
                to_float(row.get("macd_signal_ratio")),
                avg_dollar_volume_val,
                fmt_pct_value(row.get("avg_dollar_volume_pct_5")),
                parse_mmddyyyy(row.get("last_earnings_date")),
                parse_mmddyyyy(row.get("next_earnings_date")),
            ]
        )
        return output_row

    if run_mode == "single":
        data_row = build_output_row(results[0], None)
        # The section heading already carries the date, and build_output_row emits
        # no Run Date cell, so drop that column here or every value lands one
        # column left of its header.
        append_single_ticker_section(ws, headline, header_row[1:], descriptors[1:], data_row)
        ipo_start = date.today()
        ipo_end = ipo_start + timedelta(days=60)
        session = requests.Session()
        ipo_rows = fetch_alphavantage_upcoming_ipos(args.alphavantage_api_key, session, ipo_start, ipo_end)
        write_upcoming_ipos_sheet(wb, ipo_rows, ipo_start, ipo_end)
        earnings_start = date.today()
        earnings_end = earnings_start + timedelta(days=14)
        earnings_rows = fetch_nasdaq_upcoming_earnings(session, earnings_start, earnings_end)
        write_upcoming_earnings_sheet(wb, earnings_rows, earnings_start, earnings_end, qualified_dates)
        write_how_it_works_sheet(wb, args, avg_vol_mode)
        auto_size_columns(ws)
        wb.save(out_path)
        print(f"Wrote single ticker to {out_path} (Single Tickers)")
        return

    def daily_result_sort_key(row: dict[str, Any]) -> tuple[bool, int, str]:
        rank = _coerce_int(row.get("rank")) if args.score_enable else None
        daily_order = _coerce_int(row.get("_daily_order")) or 0
        symbol = str(row.get("symbol", "")).strip().upper()
        return (rank is None, rank if rank is not None else daily_order, symbol)

    new_run_rows: list[list[Any]] = []
    for row in sorted(daily_output_rows, key=daily_result_sort_key):
        symbol_display = str(row.get("symbol", "")).strip()
        symbol_key = symbol_display.upper()
        prev_count = int(prev_counts.get(symbol_key, 0))
        new_run_rows.append([data_date] + build_output_row(row, prev_count))

    for row_idx in range(ws.max_row, 2, -1):
        if _coerce_date(ws.cell(row=row_idx, column=1).value) == data_date:
            ws.delete_rows(row_idx)
    if new_run_rows:
        has_older_run = any(
            _coerce_date(ws.cell(row=row_idx, column=1).value) is not None
            for row_idx in range(3, ws.max_row + 1)
        )
        separator_already_at_top = has_older_run and all(
            ws.cell(row=3, column=col_idx).value is None
            for col_idx in range(1, ws.max_column + 1)
        )
        rows_to_insert = len(new_run_rows) + (1 if has_older_run and not separator_already_at_top else 0)
        ws.insert_rows(3, amount=rows_to_insert)
        for row_offset, values in enumerate(new_run_rows, start=3):
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row=row_offset, column=col_idx, value=value)

    black = Color(indexed=8)
    white = Color(indexed=9)
    red = Color(indexed=10)
    blue = Color(indexed=12)
    yellow = Color(indexed=13)

    header_fill = PatternFill(fill_type="solid", fgColor=black)
    descriptor_fill = PatternFill(fill_type="solid", fgColor=blue)
    white_fill = PatternFill(fill_type="solid", fgColor=white)
    black_fill = PatternFill(fill_type="solid", fgColor=black)

    header_font = Font(name="Calibri", size=13, bold=True, color=white)
    descriptor_font = Font(name="Calibri", size=11, bold=True, italic=True, color=black)
    symbol_font = Font(name="Calibri", size=13, bold=True, color=white)
    bold_font = Font(name="Calibri", size=11, bold=True, color=black)

    # Headers wrap so the spelled-out names fit their existing column widths.
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    descriptor_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_center_align = Alignment(horizontal="left", vertical="center")
    center_bottom_align = Alignment(horizontal="center", vertical="bottom")
    bottom_align = Alignment(vertical="bottom")

    thin_red = Side(style="thin", color=red)
    thin_black = Side(style="thin", color=black)
    thick_black = Side(style="thick", color=black)

    score_shift = 1 if args.score_enable else 0

    def shift_col_idx(idx: int) -> int:
        return idx + 1 + (score_shift if idx > 2 else 0)

    def shift_cols(cols: set[int]) -> set[int]:
        return {shift_col_idx(col) for col in cols}

    header_style_cols = shift_cols({1, 2, 3, 4, 6, 8, 10, 12, 14, 16, 17, 19, 21, 23, 25})
    header_value_cols = shift_cols({2, 3, 4, 6, 8, 10, 12, 14, 16, 17, 19, 21, 23, 25})
    header_style_cols.add(1)
    if args.score_enable:
        header_style_cols.update({4})
        header_value_cols.update({4})
    left_thick_cols = shift_cols({4, 6, 8, 10, 12, 14, 16, 17, 19, 21, 23, 25})
    right_thick_cols = shift_cols({5, 7, 9, 11, 13, 15, 16, 18, 20, 22, 24, 26})
    header_right_cols = {shift_col_idx(25), shift_col_idx(26)}

    # Column widths from the sample spreadsheet
    base_column_widths = {
        1: 8.0,
        2: 12.0,
        3: 11.0,
        4: 9.0,
        5: 12.0,
        6: 13.5,
        7: 8.44531,
        8: 13.1562,
        9: 9.39844,
        10: 10.6641,
        11: 13.0,
        12: 10.0,
        13: 11.0,
        14: 14.0,
        15: 10.0,
        16: 17.0,
        17: 15.0,
        18: 11.0,
        19: 9.0,
        20: 13.0,
        21: 12.2812,
        22: 14.7578,
        23: 15.6797,
        24: 12.5312,
        25: 14.2734,
        26: 13.0,
    }
    column_widths = (
        {shift_col_idx(k): v for k, v in base_column_widths.items()}
    )
    column_widths[1] = 13.0
    if args.score_enable:
        column_widths[4] = 6.0
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 50.85
    ws.row_dimensions[2].height = 40.0
    for row_idx in range(3, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 17.0

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        if col_idx in header_style_cols:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        if col_idx in header_value_cols:
            cell.number_format = "@"
        cell.border = Border(
            left=thin_red if col_idx == 1 else None,
            right=thin_red if col_idx in header_right_cols else None,
            top=thin_red,
        )

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = white_fill
        cell.font = descriptor_font
        cell.alignment = descriptor_align
        if cell.value is not None:
            cell.number_format = "@"
        cell.border = Border(
            left=thin_red if col_idx == 1 else (thick_black if col_idx in left_thick_cols else None),
            right=thick_black if col_idx in right_thick_cols else None,
        )

    bold_cols = shift_cols({4, 6, 8})
    left_center_cols = shift_cols({7, 9, 11})
    center_bottom_cols = shift_cols({17, 19, 21})
    # Percentage-like values are already computed as percent points (e.g. 21.59),
    # so use a literal percent-sign format to avoid Excel multiplying by 100.
    pct_literal_format = '0.00"%"'
    base_number_formats = {
        1: "@",
        2: "@",
        3: "0",
        4: '"$"#,##0.00',
        5: pct_literal_format,
        6: '"$"#,##0.00',
        7: "General",
        8: '"$"#,##0.00',
        9: "General",
        10: "mmm d, yyyy",
        11: "General",
        12: "0.00",
        13: pct_literal_format,
        14: "0.00",
        15: pct_literal_format,
        16: pct_literal_format,
        17: "0.00",
        18: pct_literal_format,
        19: "0.00",
        20: pct_literal_format,
        21: "0.00",
        22: pct_literal_format,
        23: '"$"#,##0.00',
        24: pct_literal_format,
        25: "mmm d, yyyy",
        26: "mmm d, yyyy",
    }
    number_formats = (
        {shift_col_idx(k): v for k, v in base_number_formats.items()}
    )
    number_formats[1] = "mmm d, yyyy"
    if args.score_enable:
        number_formats[4] = "0"

    for row_idx in range(3, ws.max_row + 1):
        is_separator = all(
            ws.cell(row=row_idx, column=col_idx).value is None
            for col_idx in range(1, ws.max_column + 1)
        )
        if is_separator:
            ws.row_dimensions[row_idx].height = 7.0
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = descriptor_fill
                cell.border = Border()
            continue
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = black_fill if col_idx == 2 else white_fill
            if col_idx == 2:
                cell.font = symbol_font
            elif col_idx in bold_cols:
                cell.font = bold_font
            if col_idx in left_center_cols:
                cell.alignment = left_center_align
            elif col_idx in center_bottom_cols:
                cell.alignment = center_bottom_align
            else:
                cell.alignment = bottom_align
            if col_idx in number_formats:
                cell.number_format = number_formats[col_idx]
            cell.border = Border(
                left=thick_black if col_idx in left_thick_cols else thin_black,
                right=thick_black if col_idx in right_thick_cols else thin_black,
            )

    merge_pairs = [
        (4, 5),
        (6, 7),
        (8, 9),
        (10, 11),
        (12, 13),
        (14, 15),
        (17, 18),
        (19, 20),
        (21, 22),
        (23, 24),
        (25, 26),
    ]
    if args.score_enable:
        merge_pairs = [(shift_col_idx(start), shift_col_idx(end)) for start, end in merge_pairs]
    for start_col, end_col in merge_pairs:
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

    ws.freeze_panes = None

    ipo_start = date.today()
    ipo_end = ipo_start + timedelta(days=60)
    session = requests.Session()
    ipo_rows = fetch_alphavantage_upcoming_ipos(args.alphavantage_api_key, session, ipo_start, ipo_end)
    write_upcoming_ipos_sheet(wb, ipo_rows, ipo_start, ipo_end)
    earnings_start = date.today()
    earnings_end = earnings_start + timedelta(days=14)
    earnings_rows = fetch_nasdaq_upcoming_earnings(session, earnings_start, earnings_end)
    write_upcoming_earnings_sheet(wb, earnings_rows, earnings_start, earnings_end, qualified_dates)
    remove_inactive_report_sheets(wb)

    # Temporarily disabled: keep only the Simulation tab in generated results.
    # Re-enable this block to restore the Top 10 OHLC Tracking tab.
    # write_top10_ohlc_tracking_sheet(wb, symbol_paths, root, top_n=10)

    # Temporarily disabled: keep only the Simulation tab in generated results.
    # Re-enable this block to restore the full Investment Dashboard tab.
    # write_investment_dashboard_sheet(
    #     wb,
    #     symbol_paths,
    #     root,
    #     top_n=10,
    #     polygon_api_key=args.polygon_api_key.strip(),
    #     intraday_exit_source=args.intraday_exit_source,
    #     market_regime_mode=args.market_regime_mode,
    # )

    write_summary_only_sheet(
        wb,
        symbol_paths,
        root,
        top_n=10,
        polygon_api_key=args.polygon_api_key.strip(),
        intraday_exit_source=args.intraday_exit_source,
        market_regime_mode=args.market_regime_mode,
    )
    write_how_it_works_sheet(wb, args, avg_vol_mode)
    prune_old_run_sheets(wb, keep_runs=15)

    wb.save(out_path)

    print(f"Wrote {len(daily_output_rows)} of {len(results)} matches to {out_path} ({sheet_name})")


if __name__ == "__main__":
    main()
